#!/usr/bin/env python3
"""Genera una entrega academica de metrados y presupuesto."""

from __future__ import annotations

import json
import math
import shutil
import subprocess
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Tuple

from .conversor_unidades import calcular_compra_comercial, normalizar_unidad
from .modelos import ComparativaItem, MaterialBOM, ResultadoProveedor


GRUPOS_PROVEEDOR = {
    "Promart": {"promart"},
    "Sodimac/Maestro": {"sodimac", "maestro"},
    "Mercado Libre Peru": {"mercadolibre"},
}
TIPOS_ESTIMADOS = {
    "estimado_por_equivalente",
    "estimado_por_promedio",
    "estimado_por_referencia_manual",
    "pendiente_revision_docente",
}


def slug_resultado(resultado: ResultadoProveedor) -> str:
    return resultado.proveedor_slug or resultado.proveedor.lower().replace(" ", "")


def clave_material(material: MaterialBOM) -> str:
    nombre = material.nombre_original.lower()
    uso = "tierra" if "tierra" in nombre or "verde/amarillo" in nombre else "general"
    return "|".join((material.nombre_normalizado, material.categoria, normalizar_unidad(material.unidad), uso))


def _mejor_resultado(resultados: Iterable[ResultadoProveedor]) -> ResultadoProveedor | None:
    candidatos = list(resultados)
    if not candidatos:
        return None
    return min(
        candidatos,
        key=lambda r: (
            float(r.subtotal_compra_real if r.subtotal_compra_real is not None else float("inf")),
            -float(r.score_tecnico),
            -float(r.score_total),
        ),
    )


def _fila_real(material: MaterialBOM, resultado: ResultadoProveedor, grupo: str) -> Dict[str, Any]:
    return {
        "clave": clave_material(material),
        "material_id": material.id,
        "material": material.nombre_original,
        "material_normalizado": material.nombre_normalizado,
        "categoria": material.categoria,
        "cantidad_requerida": material.cantidad,
        "unidad_diseno": material.unidad,
        "proveedor_grupo": grupo,
        "proveedor": resultado.proveedor,
        "producto": resultado.producto or material.nombre_original,
        "unidad_comercial": resultado.unidad_comercial or "unidad",
        "contenido_comercial": float(resultado.contenido_comercial or 1.0),
        "unidad_contenido": resultado.unidad_contenido or material.unidad,
        "cantidad_comercial_a_comprar": float(resultado.cantidad_comercial_a_comprar or material.cantidad),
        "precio_comercial": float(resultado.precio or 0.0),
        "precio_equivalente": float(resultado.precio_equivalente_unitario or 0.0),
        "subtotal_real": float(resultado.subtotal_compra_real or 0.0),
        "sobrante": float(resultado.sobrante or 0.0),
        "tipo_precio": "verificado_tienda",
        "tipo_coincidencia": resultado.tipo_coincidencia,
        "url": resultado.url or "Sin URL directa; estimado referencial",
        "fecha": resultado.fecha_consulta or datetime.now().date().isoformat(),
        "confianza": float(resultado.score_total),
        "observacion_precio": resultado.observacion_precio or "Precio verificado en tienda con URL.",
        "observacion": "; ".join(
            texto for texto in (
                resultado.observaciones,
                resultado.observacion_conversion,
                resultado.motivo_fallo,
            ) if texto
        ),
    }


def _contenido_estimado(material: MaterialBOM) -> Tuple[str, float, str]:
    unidad = normalizar_unidad(material.unidad)
    if material.categoria == "ductos/tuberias" and unidad == "m":
        return "tubo", 3.0, "m"
    if unidad == "m":
        return "metro", 1.0, "m"
    return "unidad", 1.0, material.unidad


def _fila_estimada(
    material: MaterialBOM,
    grupo: str,
    fuentes: List[ResultadoProveedor],
    fuente_local: ResultadoProveedor | None,
) -> Dict[str, Any]:
    if fuente_local is not None:
        fuentes_utiles = [fuente_local]
        tipo = "estimado_por_equivalente"
        observacion = "Precio estimado a partir de un producto del mismo proveedor que requiere revision tecnica."
    else:
        fuentes_utiles = [r for r in fuentes if r.precio_equivalente_unitario is not None]
        tipo = "estimado_por_promedio" if len(fuentes_utiles) >= 2 else "estimado_por_equivalente"
        observacion = (
            "Precio estimado usando promedio de proveedores con evidencia directa."
            if len(fuentes_utiles) >= 2
            else "Precio estimado usando como equivalente la oferta verificada de otro proveedor."
        )

    if fuentes_utiles:
        representante = min(fuentes_utiles, key=lambda r: float(r.precio_equivalente_unitario or float("inf")))
        promedio_equivalente = sum(float(r.precio_equivalente_unitario or 0.0) for r in fuentes_utiles) / len(fuentes_utiles)
        unidad_comercial = representante.unidad_comercial or "unidad"
        contenido = float(representante.contenido_comercial or 1.0)
        unidad_contenido = representante.unidad_contenido or material.unidad
        precio_comercial = promedio_equivalente * contenido
        cantidad_comercial, subtotal, sobrante, precio_equivalente = calcular_compra_comercial(
            material.cantidad,
            material.unidad,
            precio_comercial,
            unidad_comercial,
            contenido,
        )
        origenes = ", ".join(sorted({r.proveedor for r in fuentes_utiles}))
        producto = f"Referencia equivalente: {representante.producto or material.nombre_original}"
        url = representante.url or "Sin URL directa; estimado referencial"
        observacion = f"{observacion} Fuente base: {origenes}."
        confianza = max(float(r.score_total) for r in fuentes_utiles)
        coincidencia = "similar_pendiente_revision"
    else:
        if material.precio_referencial is None or float(material.precio_referencial) <= 0:
            raise ValueError(f"El material {material.id} no tiene precio real ni referencia numerica.")
        unidad_comercial, contenido, unidad_contenido = _contenido_estimado(material)
        precio_comercial = float(material.precio_referencial) * contenido
        cantidad_comercial, subtotal, sobrante, precio_equivalente = calcular_compra_comercial(
            material.cantidad,
            material.unidad,
            precio_comercial,
            unidad_comercial,
            contenido,
        )
        tipo = "estimado_por_referencia_manual"
        producto = f"Referencia tecnica de presupuesto: {material.nombre_original}"
        url = "Sin URL directa; estimado referencial"
        observacion = "Precio referencial del BOM asignado para cerrar el presupuesto academico; requiere verificacion antes de compra."
        confianza = 0.25
        coincidencia = "similar_pendiente_revision"

    return {
        "clave": clave_material(material),
        "material_id": material.id,
        "material": material.nombre_original,
        "material_normalizado": material.nombre_normalizado,
        "categoria": material.categoria,
        "cantidad_requerida": material.cantidad,
        "unidad_diseno": material.unidad,
        "proveedor_grupo": grupo,
        "proveedor": grupo,
        "producto": producto,
        "unidad_comercial": unidad_comercial,
        "contenido_comercial": contenido,
        "unidad_contenido": unidad_contenido,
        "cantidad_comercial_a_comprar": cantidad_comercial,
        "precio_comercial": round(precio_comercial, 4),
        "precio_equivalente": round(precio_equivalente, 4),
        "subtotal_real": round(subtotal, 2),
        "sobrante": round(sobrante, 4),
        "tipo_precio": tipo,
        "tipo_coincidencia": coincidencia,
        "url": url,
        "fecha": datetime.now().date().isoformat(),
        "confianza": confianza,
        "observacion_precio": observacion,
        "observacion": "ESTIMADO: no corresponde a una cotizacion comercial directa del proveedor indicado.",
    }


def construir_filas_academicas(
    comparativas: List[ComparativaItem],
    aceptable: Callable[[ResultadoProveedor], bool],
) -> Dict[str, List[Dict[str, Any]]]:
    filas: Dict[str, List[Dict[str, Any]]] = {grupo: [] for grupo in GRUPOS_PROVEEDOR}
    for comparativa in comparativas:
        material = comparativa.material_bom
        verificadas_globales = [r for r in comparativa.resultados_por_proveedor if aceptable(r)]
        for grupo, slugs in GRUPOS_PROVEEDOR.items():
            verificadas = [r for r in verificadas_globales if slug_resultado(r) in slugs]
            mejor = _mejor_resultado(verificadas)
            if mejor is not None:
                filas[grupo].append(_fila_real(material, mejor, grupo))
                continue
            equivalentes_locales = [
                r for r in comparativa.resultados_por_proveedor
                if slug_resultado(r) in slugs
                and r.precio is not None
                and r.url
                and r.confianza_conversion >= 0.5
                and r.tipo_coincidencia != "similar_pendiente_revision"
            ]
            filas[grupo].append(
                _fila_estimada(material, grupo, verificadas_globales, _mejor_resultado(equivalentes_locales))
            )
    return filas


def consolidar_filas(filas: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    grupos: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for fila in filas:
        grupos[fila["clave"]].append(fila)
    salida: List[Dict[str, Any]] = []
    prioridad = {
        "verificado_tienda": 0,
        "estimado_por_equivalente": 1,
        "estimado_por_promedio": 2,
        "estimado_por_referencia_manual": 3,
        "pendiente_revision_docente": 4,
    }
    for clave, partidas in grupos.items():
        cantidad = sum(float(f["cantidad_requerida"]) for f in partidas)
        representante = min(
            partidas,
            key=lambda f: (prioridad.get(f["tipo_precio"], 9), float(f["precio_equivalente"])),
        )
        contenido = float(representante["contenido_comercial"] or 1.0)
        cantidad_comercial, subtotal, sobrante, precio_equivalente = calcular_compra_comercial(
            cantidad,
            representante["unidad_diseno"],
            float(representante["precio_comercial"]),
            representante["unidad_comercial"],
            contenido,
        )
        fila = dict(representante)
        fila.update({
            "clave": clave,
            "material_id": ", ".join(f["material_id"] for f in partidas),
            "material": representante["material_normalizado"],
            "partidas_metrado": len(partidas),
            "cantidad_requerida": cantidad,
            "cantidad_comercial_a_comprar": cantidad_comercial,
            "precio_equivalente": round(precio_equivalente, 4),
            "subtotal_real": round(subtotal, 2),
            "sobrante": round(sobrante, 4),
        })
        salida.append(fila)
    return sorted(salida, key=lambda f: (f["categoria"], f["material"]))


def construir_recomendado(planes: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    por_clave: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for filas in planes.values():
        for fila in filas:
            por_clave[fila["clave"]].append(fila)
    salida = []
    for opciones in por_clave.values():
        menor = min(opciones, key=lambda f: float(f["subtotal_real"]))
        verificadas = [f for f in opciones if f["tipo_precio"] == "verificado_tienda"]
        mejor_verificada = min(verificadas, key=lambda f: float(f["subtotal_real"])) if verificadas else None
        if mejor_verificada and float(mejor_verificada["subtotal_real"]) <= float(menor["subtotal_real"]) * 1.15:
            elegida = mejor_verificada
            criterio = "Oferta verificada priorizada; diferencia frente al menor valor no supera 15 %."
        else:
            elegida = menor
            criterio = (
                "Menor oferta verificada."
                if elegida["tipo_precio"] == "verificado_tienda"
                else "Menor valor referencial; requiere confirmar precio antes del pedido."
            )
        fila = dict(elegida)
        fila["mejor_proveedor"] = elegida["proveedor_grupo"]
        fila["criterio_eleccion"] = criterio
        salida.append(fila)
    return sorted(salida, key=lambda f: (f["categoria"], f["material"]))


def _items_raw(raw_bom: Dict[str, Any]) -> List[Dict[str, Any]]:
    filas = []
    for indice, item in enumerate(raw_bom.get("materiales", []), 1):
        cantidad = float(item.get("cantidad") or 0.0)
        referencia = float(item.get("precio_unit_soles") or 0.0)
        filas.append({
            "codigo": item.get("codigo") or f"M{indice:03d}",
            "descripcion_tecnica": item.get("item") or item.get("descripcion"),
            "categoria": item.get("categoria", ""),
            "unidad_diseno": item.get("unidad", "und"),
            "cantidad": cantidad,
            "circuito": item.get("circuito", ""),
            "observacion": item.get("observacion", ""),
            "fuente_metrado": item.get("fuente_metrado", "BOM"),
            "precio_referencial_unitario": referencia,
            "subtotal_referencial": round(cantidad * referencia, 2),
        })
    return filas


def _resumen_planes(planes: Dict[str, List[Dict[str, Any]]]) -> List[Dict[str, Any]]:
    filas = []
    for proveedor, plan in planes.items():
        verificadas = [f for f in plan if f["tipo_precio"] == "verificado_tienda"]
        estimadas = [f for f in plan if f["tipo_precio"] != "verificado_tienda"]
        filas.append({
            "proveedor": proveedor,
            "total_verificado": round(sum(float(f["subtotal_real"]) for f in verificadas), 2),
            "total_estimado": round(sum(float(f["subtotal_real"]) for f in estimadas), 2),
            "total_general": round(sum(float(f["subtotal_real"]) for f in plan), 2),
            "items_verificados": len(verificadas),
            "items_estimados": len(estimadas),
            "cobertura_verificada_pct": round(len(verificadas) / len(plan) * 100.0, 2) if plan else 0.0,
            "advertencia": "Los valores estimados no constituyen oferta comercial del proveedor.",
        })
    return filas


def _tabla_cruzada(planes: Dict[str, List[Dict[str, Any]]], recomendado: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    mapas = {nombre: {f["clave"]: f for f in filas} for nombre, filas in planes.items()}
    rec = {f["clave"]: f for f in recomendado}
    claves = sorted(set().union(*(set(m) for m in mapas.values())))
    filas = []
    for clave in claves:
        base = next(m[clave] for m in mapas.values() if clave in m)
        fila = {
            "material": base["material"],
            "cantidad": base["cantidad_requerida"],
            "unidad": base["unidad_diseno"],
        }
        for nombre, columna in (
            ("Promart", "promart"),
            ("Sodimac/Maestro", "sodimac_maestro"),
            ("Mercado Libre Peru", "mercado_libre"),
        ):
            dato = mapas[nombre][clave]
            fila[columna] = dato["subtotal_real"]
            fila[f"tipo_{columna}"] = dato["tipo_precio"]
        elegida = rec[clave]
        fila.update({
            "mejor_opcion": elegida["mejor_proveedor"],
            "subtotal_recomendado": elegida["subtotal_real"],
            "tipo_precio": elegida["tipo_precio"],
            "observacion": elegida["criterio_eleccion"],
        })
        filas.append(fila)
    return filas


def _tex(texto: Any) -> str:
    valor = str(texto if texto is not None else "")
    for origen, destino in (
        ("\\", r"\textbackslash{}"), ("&", r"\&"), ("%", r"\%"),
        ("$", r"\$"), ("#", r"\#"), ("_", r"\_"),
        ("{", r"\{"), ("}", r"\}"), ("~", r"\textasciitilde{}"),
        ("^", r"\textasciicircum{}"),
    ):
        valor = valor.replace(origen, destino)
    return valor


def _fmt(valor: Any) -> str:
    return f"{float(valor):,.2f}"


def escribir_metrados(output_dir: Path, raw_bom: Dict[str, Any]) -> Dict[str, Path]:
    filas = _items_raw(raw_bom)
    md = output_dir / "metrados_finales.md"
    tex = output_dir / "metrados_finales.tex"
    xlsx = output_dir / "metrados_finales.xlsx"
    lineas = [
        "# Metrados finales de materiales electricos",
        "",
        f"Partidas: **{len(filas)}**.",
        "",
        "| Codigo | Descripcion tecnica | Categoria | Unidad | Cantidad | Circuito | Fuente | Observacion |",
        "|---|---|---|---:|---:|---|---|---|",
    ]
    for f in filas:
        lineas.append(
            f"| {f['codigo']} | {f['descripcion_tecnica']} | {f['categoria']} | {f['unidad_diseno']} | "
            f"{f['cantidad']:g} | {f['circuito']} | {f['fuente_metrado']} | {f['observacion']} |"
        )
    md.write_text("\n".join(lineas) + "\n", encoding="utf-8")

    cuerpo = []
    for f in filas:
        cuerpo.append(
            f"{_tex(f['codigo'])} & {_tex(f['descripcion_tecnica'])} & {_tex(f['unidad_diseno'])} & "
            f"{f['cantidad']:g} & {_tex(f['circuito'])} & {_tex(f['fuente_metrado'])} \\\\"
        )
    tex.write_text(
        "% Requiere longtable, booktabs y array.\n"
        "\\section{Metrados de materiales electricos}\n"
        "\\scriptsize\n"
        "\\begin{longtable}{p{1.5cm}p{5.5cm}p{0.8cm}r p{1.2cm}p{4.0cm}}\n"
        "\\toprule Codigo & Descripcion & Und. & Cant. & Circuito & Fuente \\\\ \\midrule\n"
        "\\endfirsthead\\toprule Codigo & Descripcion & Und. & Cant. & Circuito & Fuente \\\\ \\midrule\\endhead\n"
        + "\n".join(cuerpo)
        + "\n\\bottomrule\\end{longtable}\n\\normalsize\n",
        encoding="utf-8",
    )
    _crear_excel_simple(xlsx, "Metrados_finales", filas)
    return {"metrados_md": md, "metrados_tex": tex, "metrados_xlsx": xlsx}


def _crear_excel_simple(path: Path, titulo: str, filas: List[Dict[str, Any]]) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo
    if filas:
        headers = list(filas[0])
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for fila in filas:
            ws.append([fila.get(h) for h in headers])
        ws.auto_filter.ref = ws.dimensions
        ws.freeze_panes = "A2"
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = min(55, max(12, max(len(str(c.value or "")) for c in col) + 2))
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def escribir_excel_final(
    path: Path,
    raw_bom: Dict[str, Any],
    filas_item: Dict[str, List[Dict[str, Any]]],
    planes: Dict[str, List[Dict[str, Any]]],
    recomendado: List[Dict[str, Any]],
    resumen: List[Dict[str, Any]],
) -> None:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    azul = PatternFill("solid", fgColor="1F4E78")
    celeste = PatternFill("solid", fgColor="D9EAF7")

    portada = wb.create_sheet("Portada")
    portada.column_dimensions["A"].width = 24
    portada.column_dimensions["B"].width = 75
    datos_portada = [
        ("INSTITUCION", "Universidad Nacional del Altiplano"),
        ("ESCUELA PROFESIONAL", "Ingenieria Mecanica Electrica"),
        ("PROYECTO", str(raw_bom.get("proyecto", "Proyecto de instalaciones electricas"))),
        ("ESTUDIANTE", str(raw_bom.get("propietario", "Por completar"))),
        ("FECHA DE GENERACION", datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("NOTA", "Presupuesto academico referencial con precios verificados y estimados."),
    ]
    portada.merge_cells("A1:B2")
    portada["A1"] = "PRESUPUESTO FINAL DE INSTALACIONES ELECTRICAS"
    portada["A1"].font = Font(size=18, bold=True, color="FFFFFF")
    portada["A1"].fill = azul
    portada["A1"].alignment = Alignment(horizontal="center", vertical="center")
    for fila, (etiqueta, valor) in enumerate(datos_portada, 4):
        portada.cell(fila, 1, etiqueta).font = Font(bold=True)
        portada.cell(fila, 1).fill = celeste
        portada.cell(fila, 2, valor)
        portada.cell(fila, 2).alignment = Alignment(wrap_text=True)

    def add_sheet(nombre: str, filas: List[Dict[str, Any]], total: bool = False) -> None:
        ws = wb.create_sheet(nombre)
        if not filas:
            ws.append(["Sin datos"])
            return
        headers = [h for h in filas[0].keys() if h != "clave"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = azul
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
        for fila in filas:
            ws.append([fila.get(h) for h in headers])
        if total and "subtotal_real" in headers:
            col = headers.index("subtotal_real") + 1
            fin = ws.max_row
            ws.cell(fin + 2, col - 1, "TOTAL").font = Font(bold=True)
            ws.cell(fin + 2, col, f"=SUM({ws.cell(2, col).coordinate}:{ws.cell(fin, col).coordinate})").font = Font(bold=True)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:{ws.cell(ws.max_row, ws.max_column).coordinate}"
        for indice, header in enumerate(headers, 1):
            if any(x in header for x in ("precio", "subtotal", "total_")):
                for fila_excel in range(2, ws.max_row + 1):
                    ws.cell(fila_excel, indice).number_format = '"S/ "#,##0.00'
        for col in ws.columns:
            ancho = min(48, max(11, max(len(str(c.value or "")) for c in col) + 2))
            ws.column_dimensions[col[0].column_letter].width = ancho

    add_sheet("Metrados_finales", _items_raw(raw_bom))
    add_sheet("Promart", planes["Promart"], total=True)
    add_sheet("Sodimac_Maestro", planes["Sodimac/Maestro"], total=True)
    add_sheet("MercadoLibre", planes["Mercado Libre Peru"], total=True)
    add_sheet("Comparativa_cruzada", _tabla_cruzada(planes, recomendado))
    add_sheet("Presupuesto_recomendado", recomendado, total=True)
    add_sheet("Resumen_por_proveedor", resumen)
    estimados = [f for filas in planes.values() for f in filas if f["tipo_precio"] != "verificado_tienda"]
    add_sheet("Precios_estimados", estimados)
    evidencias = [
        {
            "material_id": f["material_id"], "material": f["material"], "proveedor": grupo,
            "producto": f["producto"], "precio": f["precio_comercial"], "tipo_precio": f["tipo_precio"],
            "url": f["url"], "fecha": f["fecha"], "observacion": f["observacion_precio"],
        }
        for grupo, filas in filas_item.items() for f in filas
    ]
    add_sheet("Evidencias_links", evidencias)
    pendientes = [
        {
            "material": f["material"], "proveedor": f["mejor_proveedor"], "precio_asignado": f["subtotal_real"],
            "tipo_precio": f["tipo_precio"], "accion": "Verificar precio y especificacion antes de compra.",
            "observacion": f["observacion_precio"],
        }
        for f in recomendado if f["tipo_precio"] != "verificado_tienda"
    ]
    add_sheet("Pendientes_revision", pendientes)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def escribir_evidencias(path: Path, filas_item: Dict[str, List[Dict[str, Any]]]) -> None:
    lineas = ["# Evidencias y enlaces de precios", "", f"Fecha de generacion: {datetime.now().date().isoformat()}", ""]
    for grupo, filas in filas_item.items():
        lineas.extend([f"## {grupo}", ""])
        for f in filas:
            lineas.extend([
                f"### {f['material_id']} - {f['material']}",
                f"- Proveedor: {grupo}",
                f"- Producto: {f['producto']}",
                f"- Precio comercial: S/ {_fmt(f['precio_comercial'])}",
                f"- Tipo de precio: `{f['tipo_precio']}`",
                f"- URL: {f['url']}",
                f"- Fecha: {f['fecha']}",
                f"- Observacion: {f['observacion_precio']}",
                "",
            ])
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lineas), encoding="utf-8")


def escribir_resumen(
    path: Path,
    raw_bom: Dict[str, Any],
    planes: Dict[str, List[Dict[str, Any]]],
    recomendado: List[Dict[str, Any]],
    resumen: List[Dict[str, Any]],
) -> None:
    total_mixto = sum(float(f["subtotal_real"]) for f in recomendado)
    mas_costosos = sorted(recomendado, key=lambda f: float(f["subtotal_real"]), reverse=True)[:10]
    economico = min(resumen, key=lambda f: float(f["total_general"]))
    mayor_cobertura = max(resumen, key=lambda f: float(f["cobertura_verificada_pct"]))
    verificados = sum(1 for f in recomendado if f["tipo_precio"] == "verificado_tienda")
    estimados = len(recomendado) - verificados
    distribucion: Dict[str, float] = defaultdict(float)
    for fila in recomendado:
        distribucion[fila["mejor_proveedor"]] += float(fila["subtotal_real"])
    lineas = [
        f"# Resumen del presupuesto final - {raw_bom.get('proyecto', 'Proyecto')}", "",
        "## Totales por proveedor", "",
        "| Proveedor | Total verificado | Total estimado | Total general | Cobertura verificada |",
        "|---|---:|---:|---:|---:|",
    ]
    for fila in resumen:
        lineas.append(
            f"| {fila['proveedor']} | S/ {_fmt(fila['total_verificado'])} | S/ {_fmt(fila['total_estimado'])} | "
            f"S/ {_fmt(fila['total_general'])} | {fila['cobertura_verificada_pct']:.2f} % |"
        )
    lineas.extend([
        "", "## Presupuesto recomendado", "",
        f"- Total recomendado mixto: **S/ {_fmt(total_mixto)}**.",
        f"- Partidas de metrado: **{len(raw_bom.get('materiales', []))}**.",
        f"- Grupos de compra consolidados: **{len(recomendado)}**.",
        f"- Grupos con precio verificado: **{verificados}**.",
        f"- Grupos con precio estimado: **{estimados}**.",
        f"- Presupuesto completo mas economico por tienda: **{economico['proveedor']}**, S/ {_fmt(economico['total_general'])}.",
        f"- Mayor cobertura de precios verificados: **{mayor_cobertura['proveedor']}**, {mayor_cobertura['cobertura_verificada_pct']:.2f} %.",
        "", "## Distribucion recomendada del pedido", "",
    ])
    for proveedor, total in sorted(distribucion.items(), key=lambda x: x[1], reverse=True):
        lineas.append(f"- {proveedor}: S/ {_fmt(total)}.")
    lineas.extend(["", "## Materiales de mayor costo", ""])
    for fila in mas_costosos:
        lineas.append(f"- {fila['material']}: S/ {_fmt(fila['subtotal_real'])} ({fila['mejor_proveedor']}, {fila['tipo_precio']}).")
    lineas.extend([
        "", "## Advertencia", "",
        "Este es un presupuesto academico referencial. Los precios estimados no son ofertas comerciales y deben verificarse antes de emitir una orden de compra. La seleccion final de ITM, diferenciales, alimentadores, tomacorrientes especiales y SPAT requiere confirmacion tecnica en obra.",
    ])
    path.write_text("\n".join(lineas) + "\n", encoding="utf-8")


def _tabla_latex(titulo: str, filas: List[Dict[str, Any]]) -> str:
    cuerpo = []
    for f in filas:
        marca = "V" if f["tipo_precio"] == "verificado_tienda" else "E"
        cuerpo.append(
            f"{_tex(f['material'])} & {f['cantidad_requerida']:g} {_tex(f['unidad_diseno'])} & "
            f"{f['cantidad_comercial_a_comprar']:g} {_tex(f['unidad_comercial'])} & "
            f"{_fmt(f['precio_comercial'])} & {_fmt(f['subtotal_real'])} & {marca} \\\\"
        )
    return (
        f"\\subsection{{{_tex(titulo)}}}\n"
        "\\begin{landscape}\\scriptsize\n"
        "\\begin{longtable}{p{8.0cm}r r r r c}\n"
        "\\toprule Material & Metrado & Compra & P. comercial & Subtotal & Tipo \\\\ \\midrule\n"
        "\\endfirsthead\\toprule Material & Metrado & Compra & P. comercial & Subtotal & Tipo \\\\ \\midrule\\endhead\n"
        + "\n".join(cuerpo)
        + "\n\\bottomrule\\end{longtable}\\normalsize\\end{landscape}\n"
    )


def escribir_latex_final(
    output_dir: Path,
    raw_bom: Dict[str, Any],
    planes: Dict[str, List[Dict[str, Any]]],
    recomendado: List[Dict[str, Any]],
    resumen: List[Dict[str, Any]],
) -> Dict[str, Path]:
    fragmento = output_dir / "presupuesto_final.tex"
    wrapper = output_dir / "presupuesto_final_documento.tex"
    comparativa = _tabla_cruzada(planes, recomendado)
    metrado_rows = _items_raw(raw_bom)
    metrado_cuerpo = "\n".join(
        f"{_tex(f['codigo'])} & {_tex(f['descripcion_tecnica'])} & {f['cantidad']:g} & {_tex(f['unidad_diseno'])} & {_tex(f['circuito'])} \\\\"
        for f in metrado_rows
    )
    comparativa_cuerpo = "\n".join(
        f"{_tex(f['material'])} & {_fmt(f['promart'])} & {_fmt(f['sodimac_maestro'])} & {_fmt(f['mercado_libre'])} & "
        f"{_tex(f['mejor_opcion'])} & {_fmt(f['subtotal_recomendado'])} \\\\"
        for f in comparativa
    )
    recomendado_cuerpo = "\n".join(
        f"{_tex(f['material'])} & {_tex(f['mejor_proveedor'])} & {f['cantidad_comercial_a_comprar']:g} {_tex(f['unidad_comercial'])} & "
        f"{_fmt(f['subtotal_real'])} & {_tex(f['tipo_precio'])} \\\\"
        for f in recomendado
    )
    total_mixto = sum(float(f["subtotal_real"]) for f in recomendado)
    distribucion: Dict[str, float] = defaultdict(float)
    for f in recomendado:
        distribucion[f["mejor_proveedor"]] += float(f["subtotal_real"])
    principal = max(distribucion.items(), key=lambda x: x[1]) if distribucion else ("Sin definir", 0.0)
    estimados = [f for f in recomendado if f["tipo_precio"] != "verificado_tienda"]

    contenido = (
        "% Fragmento generado para integracion en el expediente del proyecto.\n"
        "% Requiere: booktabs, longtable, array, pdflscape, xcolor, hyperref.\n"
        "\\section{Metrados de materiales electricos}\n"
        f"Se validaron {len(metrado_rows)} partidas usando los DXF vigentes, sus JSON documentales y los calculos previos del BOM.\n"
        "\\begin{landscape}\\scriptsize\\begin{longtable}{p{1.7cm}p{9.0cm}r c p{1.8cm}}\n"
        "\\toprule Codigo & Descripcion & Cant. & Und. & Circuito \\\\ \\midrule\n"
        "\\endfirsthead\\toprule Codigo & Descripcion & Cant. & Und. & Circuito \\\\ \\midrule\\endhead\n"
        + metrado_cuerpo + "\n\\bottomrule\\end{longtable}\\normalsize\\end{landscape}\n"
        "\\section{Cotizacion por proveedor}\n"
        + _tabla_latex("Grupo 1: Promart", planes["Promart"])
        + _tabla_latex("Grupo 2: Sodimac/Maestro", planes["Sodimac/Maestro"])
        + _tabla_latex("Grupo 3: Mercado Libre Peru", planes["Mercado Libre Peru"])
        + "\\section{Comparativa economica}\n"
        "\\begin{landscape}\\scriptsize\\begin{longtable}{p{7.2cm}rrrrp{3.0cm}r}\n"
        "\\toprule Material & Promart & Sod./Mae. & M. Libre & Mejor & Proveedor & Recomendado \\\\ \\midrule\n"
        "\\endfirsthead\\toprule Material & Promart & Sod./Mae. & M. Libre & Mejor & Proveedor & Recomendado \\\\ \\midrule\\endhead\n"
        + comparativa_cuerpo + "\n\\bottomrule\\end{longtable}\\normalsize\\end{landscape}\n"
        "\\section{Presupuesto recomendado}\n"
        "\\begin{landscape}\\scriptsize\\begin{longtable}{p{8.0cm}p{3.2cm}r r p{4.0cm}}\n"
        "\\toprule Material & Proveedor & Compra & Subtotal & Tipo de precio \\\\ \\midrule\n"
        "\\endfirsthead\\toprule Material & Proveedor & Compra & Subtotal & Tipo de precio \\\\ \\midrule\\endhead\n"
        + recomendado_cuerpo + "\n\\midrule\\multicolumn{3}{r}{\\textbf{TOTAL}} & \\textbf{S/ " + _fmt(total_mixto) + "} & \\\\ \\bottomrule\\end{longtable}\\normalsize\\end{landscape}\n"
        "\\section{Observaciones sobre precios estimados}\n"
        f"El plan recomendado contiene {len(estimados)} grupos con precio estimado. La letra V identifica precio verificado y E identifica precio estimado. "
        "Las estimaciones se calcularon desde equivalentes, promedios de ofertas verificadas o la referencia tecnica del BOM; no constituyen cotizacion comercial.\n"
        "\\section{Conclusion}\n"
        f"El presupuesto recomendado mixto asciende a \\textbf{{S/ {_fmt(total_mixto)}}}. "
        f"La mayor participacion del pedido corresponde a {_tex(principal[0])} por S/ {_fmt(principal[1])}. "
        "Se recomienda efectuar un pedido dividido segun la tabla recomendada, priorizando enlaces verificados y solicitando una proforma final para los items estimados. "
        "Antes de comprar deben confirmarse alimentadores, salidas especiales, protecciones diferenciales y todos los componentes del SPAT.\n"
    )
    fragmento.write_text(contenido, encoding="utf-8")
    wrapper.write_text(
        "\\documentclass[11pt,a4paper]{article}\n"
        "\\usepackage[utf8]{inputenc}\n\\usepackage[T1]{fontenc}\n\\usepackage[spanish]{babel}\n"
        "\\usepackage{geometry,booktabs,longtable,array,pdflscape,xcolor,hyperref}\n"
        "\\geometry{margin=1.7cm}\n\\hypersetup{colorlinks=true,urlcolor=blue}\n"
        f"\\title{{Presupuesto final de instalaciones electricas\\\\{_tex(str(raw_bom.get('proyecto', 'Proyecto')))}}}\n"
        f"\\author{{{_tex(str(raw_bom.get('propietario', 'Por completar')))}}}\n"
        f"\\date{{{datetime.now().date().isoformat()}}}\n"
        "\\begin{document}\\maketitle\\tableofcontents\\clearpage\n"
        "\\input{presupuesto_final.tex}\n\\end{document}\n",
        encoding="utf-8",
    )
    pdf = output_dir / "presupuesto_final.pdf"
    log = output_dir / "presupuesto_final.latex.log"
    if shutil.which("pdflatex"):
        salida = ""
        correcto = True
        for _ in range(2):
            proceso = subprocess.run(
                ["pdflatex", "-interaction=nonstopmode", "-halt-on-error", wrapper.name],
                cwd=output_dir,
                text=True,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                check=False,
            )
            salida += proceso.stdout
            correcto = correcto and proceso.returncode == 0
            if proceso.returncode != 0:
                break
        log.write_text(salida, encoding="utf-8")
        generado = output_dir / "presupuesto_final_documento.pdf"
        if correcto and generado.exists():
            generado.replace(pdf)
    return {"latex": fragmento, "latex_documento": wrapper, "pdf": pdf, "latex_log": log}


def escribir_readme(path: Path, raw_bom: Dict[str, Any], resumen: List[Dict[str, Any]], recomendado: List[Dict[str, Any]]) -> None:
    total = sum(float(f["subtotal_real"]) for f in recomendado)
    lineas = [
        f"# Entrega final de metrados y presupuesto - {raw_bom.get('proyecto', 'Proyecto')}", "",
        "Esta carpeta contiene la entrega academica lista para revision e impresion.", "",
        f"- Total recomendado mixto: **S/ {_fmt(total)}**.",
        "- Los precios `verificado_tienda` tienen evidencia web.",
        "- Los precios `estimado_*` son referencias academicas y deben confirmarse antes de comprar.",
        "", "## Totales completos por proveedor", "",
    ]
    for fila in resumen:
        lineas.append(f"- {fila['proveedor']}: S/ {_fmt(fila['total_general'])}.")
    path.write_text("\n".join(lineas) + "\n", encoding="utf-8")


def generar_entrega_final(
    raw_bom: Dict[str, Any],
    comparativas: List[ComparativaItem],
    output_dir: Path,
    aceptable: Callable[[ResultadoProveedor], bool],
) -> Dict[str, Any]:
    output_dir.mkdir(parents=True, exist_ok=True)
    filas_item = construir_filas_academicas(comparativas, aceptable)
    planes = {grupo: consolidar_filas(filas) for grupo, filas in filas_item.items()}
    recomendado = construir_recomendado(planes)
    resumen = _resumen_planes(planes)

    paths: Dict[str, Path] = {}
    paths.update(escribir_metrados(output_dir, raw_bom))
    paths["excel_final"] = output_dir / "presupuesto_final.xlsx"
    escribir_excel_final(paths["excel_final"], raw_bom, filas_item, planes, recomendado, resumen)
    paths["resumen_final"] = output_dir / "resumen_presupuesto_final.md"
    escribir_resumen(paths["resumen_final"], raw_bom, planes, recomendado, resumen)
    paths["evidencias"] = output_dir / "evidencias" / "enlaces_productos.md"
    escribir_evidencias(paths["evidencias"], filas_item)
    paths.update(escribir_latex_final(output_dir, raw_bom, planes, recomendado, resumen))
    paths["readme"] = output_dir / "README.md"
    escribir_readme(paths["readme"], raw_bom, resumen, recomendado)
    (output_dir / "bom_final.json").write_text(
        json.dumps(raw_bom, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    data = {
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "partidas_metrado": len(raw_bom.get("materiales", [])),
        "grupos_compra": len(recomendado),
        "resumen_por_proveedor": resumen,
        "presupuesto_recomendado": recomendado,
        "total_recomendado_mixto": round(sum(float(f["subtotal_real"]) for f in recomendado), 2),
    }
    (output_dir / "presupuesto_final.json").write_text(
        json.dumps(data, indent=2, ensure_ascii=False) + "\n", encoding="utf-8"
    )
    return {"paths": paths, "data": data, "planes": planes, "filas_item": filas_item}

