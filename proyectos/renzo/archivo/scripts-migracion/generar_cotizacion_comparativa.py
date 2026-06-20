import os
import json
import subprocess
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

script_dir = os.path.dirname(os.path.abspath(__file__))
project_dir = os.path.dirname(script_dir)
base_dir = os.path.dirname(project_dir)
scratch_dir = os.path.join(project_dir, "cotizaciones")
results_path = os.path.join(scratch_dir, "resultados_renzo_general.json")
excel_path = os.path.join(scratch_dir, "cotizacion_comparativa_renzo.xlsx")
latex_path = os.path.join(scratch_dir, "cotizacion_comparativa_renzo.tex")

# Load calculations
with open(results_path, "r", encoding="utf-8") as f:
    results = json.load(f)

# Consolidated materials list for Renzo's project
# We aggregate materials based on the 7 circuits (C1 to C7) and specifications
materials_raw = [
    {
        "item": "Cable TW THW 10 mm2 (Alimentador / Tierra)",
        "brand": "Indeco",
        "norm": "NTP-IEC 60228",
        "unit": "m",
        "qty": 125, # 75m (Alimentador) + 50m (Tierra)
        "prices": {
            "sodimac": 14.80,
            "promart": 14.50,
            "mercadolibre": 16.00,
            "maestro": 15.00,
            "google": 13.90
        },
        "use": "Alimentador principal y linea de tierra"
    },
    {
        "item": "Cable TW THW 1.5 mm2 (Alumbrado)",
        "brand": "Indeco",
        "norm": "NTP-IEC 60228",
        "unit": "m",
        "qty": 150, # C1, C4, C6
        "prices": {
            "sodimac": 1.70,
            "promart": 1.65,
            "mercadolibre": 1.90,
            "maestro": 1.75,
            "google": 1.60
        },
        "use": "Circuitos de Alumbrado C1, C4, C6"
    },
    {
        "item": "Cable TW THW 2.5 mm2 (Tomacorrientes)",
        "brand": "Indeco",
        "norm": "NTP-IEC 60228",
        "unit": "m",
        "qty": 220, # C2, C3, C5, C7
        "prices": {
            "sodimac": 2.60,
            "promart": 2.55,
            "mercadolibre": 2.90,
            "maestro": 2.65,
            "google": 2.50
        },
        "use": "Circuitos de Tomacorrientes y Cocina"
    },
    {
        "item": "Tubo PVC SAP 32 mm",
        "brand": "Pavco Wavin",
        "norm": "NTP 399.006",
        "unit": "m",
        "qty": 25,
        "prices": {
            "sodimac": 6.80,
            "promart": 6.50,
            "mercadolibre": 7.50,
            "maestro": 6.70,
            "google": 6.20
        },
        "use": "Canalizacion del Alimentador principal"
    },
    {
        "item": "Tubo PVC SAP 20 mm",
        "brand": "Pavco Wavin",
        "norm": "NTP 399.006",
        "unit": "m",
        "qty": 185, # Total derived circuits
        "prices": {
            "sodimac": 2.30,
            "promart": 2.25,
            "mercadolibre": 2.80,
            "maestro": 2.35,
            "google": 2.20
        },
        "use": "Canalizaciones de circuitos derivados"
    },
    {
        "item": "ITM 2P-40A (Interruptor General)",
        "brand": "Bticino",
        "norm": "NTP-IEC 60898-1",
        "unit": "und",
        "qty": 1,
        "prices": {
            "sodimac": 45.90,
            "promart": 42.90,
            "mercadolibre": 49.00,
            "maestro": 46.50,
            "google": 39.90
        },
        "use": "Proteccion general del Tablero TG-01"
    },
    {
        "item": "ITM 2P-10A (Alumbrado)",
        "brand": "Bticino",
        "norm": "NTP-IEC 60898-1",
        "unit": "und",
        "qty": 3, # C1, C4, C6
        "prices": {
            "sodimac": 38.50,
            "promart": 39.90,
            "mercadolibre": 42.00,
            "maestro": 39.00,
            "google": 36.90
        },
        "use": "Proteccion termomagnetica C1, C4, C6"
    },
    {
        "item": "ITM 2P-16A (Tomacorrientes)",
        "brand": "Bticino",
        "norm": "NTP-IEC 60898-1",
        "unit": "und",
        "qty": 3, # C2, C5, C7
        "prices": {
            "sodimac": 39.90,
            "promart": 39.00,
            "mercadolibre": 45.00,
            "maestro": 40.50,
            "google": 37.90
        },
        "use": "Proteccion termomagnetica C2, C5, C7"
    },
    {
        "item": "ITM 2P-20A (Cocina)",
        "brand": "Bticino",
        "norm": "NTP-IEC 60898-1",
        "unit": "und",
        "qty": 1, # C3
        "prices": {
            "sodimac": 42.50,
            "promart": 39.90,
            "mercadolibre": 48.00,
            "maestro": 43.00,
            "google": 38.90
        },
        "use": "Proteccion termomagnetica C3"
    },
    {
        "item": "Interruptor diferencial 2P-40A-30mA",
        "brand": "Schneider Electric",
        "norm": "NTP-IEC 61008-1",
        "unit": "und",
        "qty": 1,
        "prices": {
            "sodimac": 160.60,
            "promart": 179.00,
            "mercadolibre": 165.00,
            "maestro": 162.00,
            "google": 159.00
        },
        "use": "Proteccion diferencial general"
    },
    {
        "item": "Interruptor diferencial 2P-25A-30mA",
        "brand": "Schneider Electric",
        "norm": "NTP-IEC 61008-1",
        "unit": "und",
        "qty": 4, # C2, C3, C5, C7
        "prices": {
            "sodimac": 148.00,
            "promart": 145.00,
            "mercadolibre": 155.00,
            "maestro": 149.00,
            "google": 139.90
        },
        "use": "Proteccion diferencial C2, C3, C5, C7"
    },
    {
        "item": "Tablero general para 8 circuitos",
        "brand": "Bticino",
        "norm": "IEC 61439",
        "unit": "und",
        "qty": 1,
        "prices": {
            "sodimac": 79.90,
            "promart": 75.00,
            "mercadolibre": 85.00,
            "maestro": 81.00,
            "google": 72.00
        },
        "use": "Distribucion de circuitos en la vivienda"
    },
    {
        "item": "Caja rectangular/octogonal FG",
        "brand": "Krosch",
        "norm": "NTP 370.054",
        "unit": "und",
        "qty": 47,
        "prices": {
            "sodimac": 3.80,
            "promart": 3.50,
            "mercadolibre": 4.50,
            "maestro": 3.90,
            "google": 3.20
        },
        "use": "Cajas de pase y salida en techo/pared"
    },
    {
        "item": "Varilla de puesta a tierra 5/8 x 2.4 m",
        "brand": "Jesa",
        "norm": "NTP 370.056",
        "unit": "und",
        "qty": 1,
        "prices": {
            "sodimac": 139.00,
            "promart": 135.00,
            "mercadolibre": 145.00,
            "maestro": 140.00,
            "google": 129.00
        },
        "use": "Electrodo de cobre para Pozo a Tierra"
    },
    {
        "item": "Cinta de aislar electrica",
        "brand": "3M",
        "norm": "ASTM D3000",
        "unit": "und",
        "qty": 5,
        "prices": {
            "sodimac": 7.80,
            "promart": 7.50,
            "mercadolibre": 9.00,
            "maestro": 8.00,
            "google": 6.90
        },
        "use": "Aislamiento de empalmes"
    }
]

# Create Excel workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Comparativa Proveedores"
ws.views.sheetView[0].showGridLines = True

# Styling
fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
fill_accent = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")
font_title = Font(name="Calibri", size=16, bold=True, color="1B365D")
font_subtitle = Font(name="Calibri", size=10, italic=True, color="4A5568")
font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
font_body = Font(name="Calibri", size=11)
font_total = Font(name="Calibri", size=11, bold=True)

thin_border = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0")
)
double_bottom = Border(
    top=Side(style="thin", color="1B365D"),
    bottom=Side(style="double", color="1B365D")
)

# Header Info
ws["A1"] = "PRESUPUESTO COMPARATIVO DE PROVEEDORES - PROYECTO RENZO"
ws["A1"].font = font_title
ws["A2"] = f"Propietario: {results['propietario']}  |  Curso: Instalaciones Eléctricas I  |  Fecha: 2026-06-08"
ws["A2"].font = font_subtitle

# Headers
headers = [
    "Item", "Descripcion del Material", "Marca", "Norma Tecnica", "Unidad", "Cant",
    "Sodimac (S/)", "Promart (S/)", "MercadoLibre (S/)", "Maestro (S/)", "Google (S/)",
    "Mejor Precio (S/)", "Mejor Proveedor", "Subtotal (S/)"
]

row_num = 4
for col_num, h in enumerate(headers, 1):
    cell = ws.cell(row=row_num, column=col_num, value=h)
    cell.font = font_header
    cell.fill = fill_header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row_num].height = 30

start_row = 5
for idx, m in enumerate(materials_raw, 1):
    row_num += 1
    ws.row_dimensions[row_num].height = 20
    is_even = (row_num % 2 == 0)
    row_fill = fill_zebra if is_even else None
    
    # Metadata
    ws.cell(row=row_num, column=1, value=idx).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num, column=2, value=m["item"]).alignment = Alignment(horizontal="left")
    ws.cell(row=row_num, column=3, value=m["brand"]).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num, column=4, value=m["norm"]).alignment = Alignment(horizontal="left")
    ws.cell(row=row_num, column=5, value=m["unit"]).alignment = Alignment(horizontal="center")
    ws.cell(row=row_num, column=6, value=m["qty"]).alignment = Alignment(horizontal="right")
    
    # Prices per supplier
    suppliers = ["sodimac", "promart", "mercadolibre", "maestro", "google"]
    for s_idx, s in enumerate(suppliers, 7):
        p_cell = ws.cell(row=row_num, column=s_idx, value=m["prices"][s])
        p_cell.number_format = '"S/" #,##0.00'
        p_cell.alignment = Alignment(horizontal="right")
        
    # Formulas for Best Price and Best Supplier
    # G is Sodimac, K is Google
    best_price_formula = f"=MIN(G{row_num}:K{row_num})"
    bp_cell = ws.cell(row=row_num, column=12, value=best_price_formula)
    bp_cell.number_format = '"S/" #,##0.00'
    bp_cell.alignment = Alignment(horizontal="right")
    
    # Better Supplier Formula based on MIN price
    best_supplier_formula = f'=IF(L{row_num}=G{row_num},"Sodimac",IF(L{row_num}=H{row_num},"Promart",IF(L{row_num}=I{row_num},"Mercado Libre",IF(L{row_num}=J{row_num},"Maestro","Google"))))'
    bs_cell = ws.cell(row=row_num, column=13, value=best_supplier_formula)
    bs_cell.alignment = Alignment(horizontal="center")
    
    # Subtotal = Quantity * Best Price
    subtotal_formula = f"=F{row_num}*L{row_num}"
    sub_cell = ws.cell(row=row_num, column=14, value=subtotal_formula)
    sub_cell.number_format = '"S/" #,##0.00'
    sub_cell.alignment = Alignment(horizontal="right")
    
    for c in range(1, 15):
        cell = ws.cell(row=row_num, column=c)
        cell.font = font_body
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill

# Totals Block
end_row = row_num
total_materials_row = end_row + 2

# Row total materials
ws.cell(row=total_materials_row, column=13, value="Total Materiales:").font = font_total
ws.cell(row=total_materials_row, column=13).alignment = Alignment(horizontal="right")
tot_m_cell = ws.cell(row=total_materials_row, column=14, value=f"=SUM(N5:N{end_row})")
tot_m_cell.font = font_total
tot_m_cell.number_format = '"S/" #,##0.00'
tot_m_cell.border = double_bottom

# Mano de obra (40% de materiales)
mo_row = total_materials_row + 1
ws.cell(row=mo_row, column=13, value="Mano de Obra (40%):").font = font_body
ws.cell(row=mo_row, column=13).alignment = Alignment(horizontal="right")
tot_mo_cell = ws.cell(row=mo_row, column=14, value=f"=N{total_materials_row}*0.40")
tot_mo_cell.font = font_body
tot_mo_cell.number_format = '"S/" #,##0.00'

# Subtotal General
sub_row = mo_row + 1
ws.cell(row=sub_row, column=13, value="Subtotal General:").font = font_body
ws.cell(row=sub_row, column=13).alignment = Alignment(horizontal="right")
tot_sub_cell = ws.cell(row=sub_row, column=14, value=f"=N{total_materials_row}+N{mo_row}")
tot_sub_cell.font = font_body
tot_sub_cell.number_format = '"S/" #,##0.00'

# IGV (18%)
igv_row = sub_row + 1
ws.cell(row=igv_row, column=13, value="IGV (18%):").font = font_body
ws.cell(row=igv_row, column=13).alignment = Alignment(horizontal="right")
tot_igv_cell = ws.cell(row=igv_row, column=14, value=f"=N{sub_row}*0.18")
tot_igv_cell.font = font_body
tot_igv_cell.number_format = '"S/" #,##0.00'

# Grand Total
gt_row = igv_row + 1
ws.cell(row=gt_row, column=13, value="TOTAL GENERAL:").font = font_total
ws.cell(row=gt_row, column=13).fill = fill_accent
ws.cell(row=gt_row, column=13).alignment = Alignment(horizontal="right")
tot_gt_cell = ws.cell(row=gt_row, column=14, value=f"=N{sub_row}+N{igv_row}")
tot_gt_cell.font = font_total
tot_gt_cell.fill = fill_accent
tot_gt_cell.number_format = '"S/" #,##0.00'
tot_gt_cell.border = Border(top=Side(style="thin", color="1B365D"), bottom=Side(style="double", color="1B365D"))

# Auto-fit Column Widths
for col in ws.columns:
    max_len = 0
    col_letter = get_column_letter(col[0].column)
    for cell in col:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 30
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 24
ws.column_dimensions['E'].width = 6
ws.column_dimensions['F'].width = 6

wb.save(excel_path)
print(f"Excel comparativo de Renzo guardado exitosamente en: {excel_path}")

# Generate LaTeX stand-alone file
print("Generating LaTeX source for Renzo's comparative report...")
latex_content = r"""\documentclass[11pt,a4paper]{article}
\usepackage[utf8]{inputenc}
\usepackage[T1]{fontenc}
\usepackage[spanish]{babel}
\usepackage{geometry}
\usepackage{booktabs}
\usepackage{xcolor}
\usepackage{fancyhdr}
\usepackage{tabularx}
\usepackage{array}
\usepackage{float}
\usepackage{pdflscape}

\geometry{left=1.5cm,right=1.5cm,top=2.5cm,bottom=2.5cm}
\definecolor{navyblue}{HTML}{1B365D}
\definecolor{lightgray}{HTML}{F7FAFC}

\pagestyle{fancy}
\fancyhf{}
\lhead{\color{navyblue}\bfseries ANÁLISIS COMPARATIVO DE PRECIOS - PROYECTO RENZO}
\rhead{\scriptsize UNAP - FIME}
\rfoot{Página \thepage}

\setlength{\parindent}{0pt}
\setlength{\parskip}{6pt}

\begin{document}

\begin{landscape}
\begin{center}
    {\LARGE\bfseries\color{navyblue} ANÁLISIS COMPARATIVO DE PRECIOS Y PROVEEDORES}\\[0.2cm]
    {\large\bfseries PROYECTO: VIVIENDA UNIFAMILIAR - RENZO GABRIEL MAMANI GALINDO}
\end{center}

\vspace{0.4cm}
\textbf{Propietario:} """ + results["propietario"] + r""" \hfill \textbf{Fecha:} 2026-06-08 \\
\textbf{Curso:} Instalaciones Eléctricas I \hfill \textbf{Agente Validador:} Antigravity AI \\

\vspace{0.4cm}

\begin{table}[H]
\centering
\begin{scriptsize}
\begin{tabularx}{1.38\textwidth}{c X c c c c c c c c c}
\toprule
\textbf{Item} & \textbf{Material} & \textbf{Cant} & \textbf{Sodimac} & \textbf{Promart} & \textbf{M.Libre} & \textbf{Maestro} & \textbf{Google} & \textbf{Mejor P.} & \textbf{Mejor Prov.} & \textbf{Subtotal} \\
\midrule
"""

total_materials_val = 0.0
for idx, m in enumerate(materials_raw, 1):
    item_esc = m["item"].replace("%", r"\%").replace("&", r"\&").replace('"', "''")
    prices_list = [m["prices"]["sodimac"], m["prices"]["promart"], m["prices"]["mercadolibre"], m["prices"]["maestro"], m["prices"]["google"]]
    best_p = min(prices_list)
    
    # Identify which supplier has the best price
    best_prov = "Google"
    if best_p == m["prices"]["sodimac"]: best_prov = "Sodimac"
    elif best_p == m["prices"]["promart"]: best_prov = "Promart"
    elif best_p == m["prices"]["mercadolibre"]: best_prov = "M.Libre"
    elif best_p == m["prices"]["maestro"]: best_prov = "Maestro"
    
    subt = best_p * m["qty"]
    total_materials_val += subt
    
    latex_content += f"{idx} & {item_esc} & {m['qty']} & S/ {m['prices']['sodimac']:.2f} & S/ {m['prices']['promart']:.2f} & S/ {m['prices']['mercadolibre']:.2f} & S/ {m['prices']['maestro']:.2f} & S/ {m['prices']['google']:.2f} & S/ {best_p:.2f} & {best_prov} & S/ {subt:.2f} \\\\\n"

mano_obra = total_materials_val * 0.40
subtotal_general = total_materials_val + mano_obra
igv = subtotal_general * 0.18
total_general = subtotal_general + igv

latex_content += r"""\midrule
\multicolumn{10}{r}{\textbf{Total Materiales:}} & S/ """ + f"{total_materials_val:,.2f}" + r""" \\
\multicolumn{10}{r}{Mano de Obra (40\%):} & S/ """ + f"{mano_obra:,.2f}" + r""" \\
\multicolumn{10}{r}{Subtotal General:} & S/ """ + f"{subtotal_general:,.2f}" + r""" \\
\multicolumn{10}{r}{IGV (18\%):} & S/ """ + f"{igv:,.2f}" + r""" \\
\multicolumn{10}{r}{\textbf{TOTAL GENERAL:}} & \textbf{S/ """ + f"{total_general:,.2f}" + r"""} \\
\bottomrule
\end{tabularx}
\end{scriptsize}
\end{table}
\end{landscape}

\end{document}
"""

with open(latex_path, "w", encoding="utf-8") as f:
    f.write(latex_content)

print(f"LaTeX de Renzo guardado en: {latex_path}")

# Compile LaTeX to PDF
print("Compiling LaTeX to PDF...")
cmd = ["pdflatex", "-interaction=nonstopmode", "cotizacion_comparativa_renzo.tex"]
result = subprocess.run(cmd, cwd=scratch_dir, capture_output=True, text=True)

if result.returncode == 0:
    print(f"SUCCESS: PDF cotizacion de Renzo generado en: {os.path.join(scratch_dir, 'cotizacion_comparativa_renzo.pdf')}")
else:
    print("Error al compilar PDF de Renzo:")
    print(result.stderr or result.stdout)

print("\n--- RESUMEN DE PROCESAMIENTO RENZO ---")
print(f"  Subtotal Materiales: S/ {total_materials_val:.2f}")
print(f"  Mano de Obra (40%):  S/ {mano_obra:.2f}")
print(f"  IGV (18%):           S/ {igv:.2f}")
print(f"  TOTAL GENERAL:       S/ {total_general:.2f}")
print("=======================================")
