import os
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
folders = ["02_memoria_calculo/hojas_excel", "04_metrados_y_presupuesto"]

for folder in folders:
    os.makedirs(os.path.join(base_dir, folder), exist_ok=True)

# Styles helper
font_title = Font(name="Arial", size=14, bold=True, color="1B365D")
font_subtitle = Font(name="Arial", size=10, italic=True, color="4A5568")
font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
font_body = Font(name="Arial", size=10)
font_total = Font(name="Arial", size=10, bold=True)

fill_header = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid")
fill_zebra = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
fill_accent = PatternFill(start_color="EBF8FF", end_color="EBF8FF", fill_type="solid")

thin_side = Side(border_style="thin", color="CBD5E0")
double_side = Side(border_style="double", color="1B365D")
border_cell = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
border_total = Border(top=thin_side, bottom=double_side)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="center")
align_right = Alignment(horizontal="right", vertical="center")

def format_ws(ws, title_text, col_widths):
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells("A1:G1")
    ws["A1"] = title_text
    ws["A1"].font = font_title
    ws["A1"].alignment = align_left
    
    ws.merge_cells("A2:G2")
    ws["A2"] = "Estudiante: Renzo Gabriel Mamani Galindo  |  Curso: Instalaciones Eléctricas I  |  Universidad: UNAP"
    ws["A2"].font = font_subtitle
    ws["A2"].alignment = align_left
    
    # Header Row is row 4
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 26
    
    # Adjust column widths
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i+1)].width = w

# ====================================================
# 1. GENERATE CUADRO DE CARGAS SPREADSHEET
# ====================================================
def generate_cuadro_cargas():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cuadro de Cargas"
    
    col_widths = [10, 30, 16, 14, 16, 12, 28, 14, 14]
    format_ws(ws, "CUADRO DE CARGAS - RESIDENCIAL", col_widths)
    
    headers = [
        "Circuito", "Descripción de Carga", "Pot. Instalada (W)", 
        "F. Demanda", "Máx. Demanda (W)", "I. Diseño (A)", 
        "Conductor / Alimentador", "ITM (A)", "ID (mA)"
    ]
    
    # Write headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    rows_data = [
        ["C1", "Alumbrado General (Lámparas LED)", 132, 1.00, "=C5*D5", "=E5/(220*0.9)", "3 x 1.5 mm2 Cu - PVC 3/4\"", "2P - 10A", "2P - 25A / 30mA"],
        ["C2", "Tomacorrientes Generales (TC)", 2520, 0.70, "=C6*D6", "=E6/(220*0.9)", "3 x 2.5 mm2 Cu - PVC 3/4\"", "2P - 16A", "2P - 25A / 30mA"],
        ["C3", "Tomacorrientes de Cocina", 1200, 0.80, "=C7*D7", "=E7/(220*0.9)", "3 x 2.5 mm2 Cu - PVC 3/4\"", "2P - 20A", "2P - 25A / 30mA"],
        ["C4", "Servicios Lavandería", 360, 0.80, "=C8*D8", "=E8/(220*0.9)", "3 x 2.5 mm2 Cu - PVC 3/4\"", "2P - 16A", "2P - 25A / 30mA"],
        ["C5", "Carga Especial (Bomba de Agua)", 750, 1.00, "=C9*D9", "=E9/(220*0.9)", "3 x 2.5 mm2 Cu - PVC 3/4\"", "2P - 16A", "2P - 25A / 30mA"]
    ]
    
    # Write rows
    for row_idx, r in enumerate(rows_data, start=5):
        ws.row_dimensions[row_idx].height = 20
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            # Format alignment and formats
            if col_idx in [1, 7, 8, 9]:
                cell.alignment = align_center if col_idx != 7 else align_left
            elif col_idx == 2:
                cell.alignment = align_left
            elif col_idx in [3, 5]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx == 4:
                cell.alignment = align_right
                cell.number_format = "0.00"
            elif col_idx == 6:
                cell.alignment = align_right
                cell.number_format = "0.00"
                
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
    # Total row
    tot_row = 10
    ws.row_dimensions[tot_row].height = 22
    tot_data = [
        "TOTAL", "DEMANDA CONSOLIDADA", "=SUM(C5:C9)", "", "=SUM(E5:E9)", "=E10/(220*0.9)",
        "Acometida: 2 x 10 mm2 Cu + PE", "Gral: 2P-40A", "ID: 2P-40A/30mA"
    ]
    
    for col_idx, val in enumerate(tot_data, start=1):
        cell = ws.cell(row=tot_row, column=col_idx, value=val)
        cell.font = font_total
        cell.border = border_total
        cell.fill = fill_accent
        
        if col_idx in [1, 7, 8, 9]:
            cell.alignment = align_center if col_idx != 7 else align_left
        elif col_idx == 2:
            cell.alignment = align_left
        elif col_idx in [3, 5]:
            cell.alignment = align_right
            cell.number_format = "#,##0"
        elif col_idx == 6:
            cell.alignment = align_right
            cell.number_format = "0.00"
            
    wb.save(os.path.join(base_dir, "02_memoria_calculo", "hojas_excel", "cuadro_cargas.xlsx"))
    print("cuadro_cargas.xlsx generado.")

# ====================================================
# 2. GENERATE MAXIMA DEMANDA SPREADSHEET
# ====================================================
def generate_maxima_demanda():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Máxima Demanda"
    
    col_widths = [8, 32, 12, 15, 12, 15, 20]
    format_ws(ws, "CÁLCULO EXPLICATIVO DE LA MÁXIMA DEMANDA", col_widths)
    
    headers = [
        "Item", "Ambiente / Concepto de Carga", "Cantidad", 
        "Pot. Unit. (W)", "Pot. Total (W)", "F. Demanda", "Máx. Demanda (W)"
    ]
    
    # Write headers
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    loads = [
        [1, "Sala-comedor - Luminaria LED", 2, 12, "=C5*D5", 1.0, "=E5*F5"],
        [2, "Sala-comedor - Tomacorriente doble", 4, 180, "=C6*D6", 0.7, "=E6*F6"],
        [3, "Cocina - Luminaria LED", 2, 12, "=C7*D7", 1.0, "=E7*F7"],
        [4, "Cocina - Pequeños Artefactos", 4, 300, "=C8*D8", 0.8, "=E8*F8"],
        [5, "Dormitorio Principal - Luminaria LED", 1, 12, "=C9*D9", 1.0, "=E9*F9"],
        [6, "Dormitorio Principal - Tomacorriente", 3, 180, "=C10*D10", 0.7, "=E10*F10"],
        [7, "Dormitorio Secundario 1 - Luz LED", 1, 12, "=C11*D11", 1.0, "=E11*F11"],
        [8, "Dormitorio Secundario 1 - Tomacorriente", 2, 180, "=C12*D12", 0.7, "=E12*F12"],
        [9, "Dormitorio Secundario 2 - Luz LED", 1, 12, "=C13*D13", 1.0, "=E13*F13"],
        [10, "Dormitorio Secundario 2 - Tomacorriente", 2, 180, "=C14*D14", 0.7, "=E14*F14"],
        [11, "Baño - Luminaria LED", 1, 12, "=C15*D15", 1.0, "=E15*F15"],
        [12, "Baño - Tomacorriente GFCI", 1, 180, "=C16*D16", 0.7, "=E16*F16"],
        [13, "Lavandería - Luminaria LED", 1, 12, "=C17*D17", 1.0, "=E17*F17"],
        [14, "Lavandería - Tomacorriente de servicio", 2, 180, "=C18*D18", 0.8, "=E18*F18"],
        [15, "Área Técnica - Bomba de Agua 1HP", 1, 750, "=C19*D19", 1.0, "=E19*F19"],
        [16, "Circulaciones (Pasadizos, Escalera) - Luz", 2, 12, "=C20*D20", 1.0, "=E20*F20"]
    ]
    
    for row_idx, r in enumerate(loads, start=5):
        ws.row_dimensions[row_idx].height = 19
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            if col_idx in [1, 3]:
                cell.alignment = align_center
            elif col_idx == 2:
                cell.alignment = align_left
            elif col_idx in [4, 5, 7]:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx == 6:
                cell.alignment = align_right
                cell.number_format = "0.00"
                
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
    # Total row
    tot_row = 21
    ws.row_dimensions[tot_row].height = 22
    tot_data = [
        "TOTAL", "CARGAS SUMADAS VIVIENDA", "=SUM(C5:C20)", "", "=SUM(E5:E20)", "", "=SUM(G5:G20)"
    ]
    
    for col_idx, val in enumerate(tot_data, start=1):
        cell = ws.cell(row=tot_row, column=col_idx, value=val)
        cell.font = font_total
        cell.border = border_total
        cell.fill = fill_accent
        
        if col_idx in [1, 3]:
            cell.alignment = align_center
        elif col_idx == 2:
            cell.alignment = align_left
        elif col_idx in [5, 7]:
            cell.alignment = align_right
            cell.number_format = "#,##0"
            
    # Technical explanation box below
    ws.merge_cells("A23:G25")
    ws["A23"] = (
        "NOTAS TÉCNICAS:\n"
        "1. Los factores de demanda son tomados según CNE - Utilización Reglas 050-200 para cargas generales.\n"
        "2. Fórmulas aplicadas: Potencia de Demanda = Potencia Instalada * Factor de Demanda.\n"
        "3. La intensidad de corriente para el alimentador monofásico es I = P / (V * cos φ) = 3,894 / (220 * 0.90) = 19.67 A."
    )
    ws["A23"].font = Font(name="Arial", size=9, color="4A5568")
    ws["A23"].alignment = Alignment(wrap_text=True, vertical="top")
    
    wb.save(os.path.join(base_dir, "02_memoria_calculo", "hojas_excel", "maxima_demanda.xlsx"))
    print("maxima_demanda.xlsx generado.")

# ====================================================
# 3. GENERATE METRADOS SPREADSHEET
# ====================================================
def generate_metrados():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Metrados Eléctricos"
    
    col_widths = [10, 12, 45, 12, 14, 25]
    format_ws(ws, "METRADOS DE INSTALACIONES ELÉCTRICAS", col_widths)
    
    headers = [
        "Item", "Código", "Descripción del Material / Partida", 
        "Unidad", "Cantidad", "Observaciones / Detalles"
    ]
    
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    metrados_data = [
        ["1.0", "CAN-01", "Tubería PVC Liviana (PV-L) 3/4\" para alumbrado", "m", 120, "Empotrado en losa de techo"],
        ["1.1", "CAN-02", "Tubería PVC Pesada (PV-P) 3/4\" para tomacorrientes", "m", 150, "Empotrado en contrapiso"],
        ["1.2", "CAN-03", "Tubería PVC Pesada (PV-P) 1\" para alimentador", "m", 25, "Desde medidor a TG y subtableros"],
        ["2.0", "CON-01", "Conductor de cobre LSOH (libre halógenos) 1.5 mm2", "m", 350, "Cableado de alumbrado (fase, neutro)"],
        ["2.1", "CON-02", "Conductor de cobre LSOH (libre halógenos) 2.5 mm2", "m", 480, "Cableado de TCs y cargas especiales (c/ tierra)"],
        ["2.2", "CON-03", "Conductor de cobre LSOH (libre halógenos) 10 mm2", "m", 30, "Alimentador general acometida"],
        ["3.0", "CAJ-01", "Cajas Octogonales de fierro galvanizado 4\" x 2\" (Salida Luz)", "pza", 12, "Para puntos de luz y derivación"],
        ["3.1", "CAJ-02", "Cajas Rectangulares de fierro galvanizado 4\" x 2\"", "pza", 28, "Para interruptores y tomacorrientes"],
        ["4.0", "TAB-01", "Tablero de Distribución Metálico TG-01 (12 polos)", "u", 1, "Tablero General Primer Piso"],
        ["4.1", "TAB-02", "Tablero de Distribución Metálico TD-01 (8 polos)", "u", 1, "Tablero Distribución Segundo Piso"],
        ["4.2", "TAB-03", "Tablero de Distribución Metálico TD-02 (8 polos)", "u", 1, "Tablero Distribución Tercer Piso"],
        ["5.0", "ACC-01", "Interruptores Simples (placas de empotrar)", "pza", 4, "Ambientes generales"],
        ["5.1", "ACC-02", "Interruptores Conmutación de 3 vías (S3)", "pza", 4, "Escaleras y pasadizo"],
        ["5.2", "ACC-03", "Placa de Tomacorriente Doble con Toma de Tierra", "pza", 22, "Tomacorrientes generales y cocina"],
        ["5.3", "ACC-04", "Placa de Tomacorriente GFCI", "pza", 2, "Protección baños (Piso 2 y Piso 3)"],
        ["6.0", "PAT-01", "Pozo de Puesta a Tierra Completo (Electrodo, gel, caja)", "jgo", 1, "Resistencia final < 15 Ohms"]
    ]
    
    for row_idx, r in enumerate(metrados_data, start=5):
        ws.row_dimensions[row_idx].height = 19
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            if col_idx in [1, 2, 4]:
                cell.alignment = align_center
            elif col_idx in [3, 6]:
                cell.alignment = align_left
            elif col_idx == 5:
                cell.alignment = align_right
                cell.number_format = "#,##0"
                
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
    wb.save(os.path.join(base_dir, "04_metrados_y_presupuesto", "metrados.xlsx"))
    print("metrados.xlsx generado.")

# ====================================================
# 4. GENERATE PRESUPUESTO SPREADSHEET
# ====================================================
def generate_presupuesto():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Presupuesto Eléctrico"
    
    col_widths = [8, 12, 45, 10, 12, 14, 16]
    format_ws(ws, "PRESUPUESTO GENERAL ESTIMADO", col_widths)
    
    headers = [
        "Item", "Código", "Descripción de Partida / Material", 
        "Unidad", "Cantidad", "P. Unitario (S/.)", "Parcial (S/.)"
    ]
    
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=h)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = border_cell
        
    items = [
        ["1.0", "CAN-01", "Tubería PVC Liviana (PV-L) 3/4\" para alumbrado", "m", 120, 2.50, "=E5*F5"],
        ["1.1", "CAN-02", "Tubería PVC Pesada (PV-P) 3/4\" para tomacorrientes", "m", 150, 4.20, "=E6*F6"],
        ["1.2", "CAN-03", "Tubería PVC Pesada (PV-P) 1\" para alimentador", "m", 25, 6.80, "=E7*F7"],
        ["2.0", "CON-01", "Conductor de cobre LSOH 1.5 mm2", "m", 350, 3.20, "=E8*F8"],
        ["2.1", "CON-02", "Conductor de cobre LSOH 2.5 mm2", "m", 480, 4.80, "=E9*F9"],
        ["2.2", "CON-03", "Conductor de cobre LSOH 10 mm2", "m", 30, 16.50, "=E10*F10"],
        ["3.0", "CAJ-01", "Cajas Octogonales FG 4\" x 2\"", "pza", 12, 3.50, "=E11*F11"],
        ["3.1", "CAJ-02", "Cajas Rectangulares FG 4\" x 2\"", "pza", 28, 3.20, "=E12*F12"],
        ["4.0", "TAB-01", "Tablero General TG-01 (12 polos) equipado completo", "u", 1, 450.00, "=E13*F13"],
        ["4.1", "TAB-02", "Tablero de Distribución TD-01 equipado completo", "u", 1, 320.00, "=E14*F14"],
        ["4.2", "TAB-03", "Tablero de Distribución TD-02 equipado completo", "u", 1, 320.00, "=E15*F15"],
        ["5.0", "ACC-01", "Placas de interruptor simple", "pza", 4, 12.00, "=E16*F16"],
        ["5.1", "ACC-02", "Placas de interruptor conmutador S3", "pza", 4, 15.00, "=E17*F17"],
        ["5.2", "ACC-03", "Placas de tomacorriente doble c/ tierra", "pza", 22, 14.50, "=E18*F18"],
        ["5.3", "ACC-04", "Placas de tomacorriente GFCI", "pza", 2, 45.00, "=E19*F19"],
        ["6.0", "PAT-01", "Kit de Pozo de Puesta a Tierra Completo c/ accesorios", "jgo", 1, 650.00, "=E20*F20"],
        ["7.0", "OBR-01", "Mano de Obra Calificada de Instalaciones Interiores", "jgo", 1, 3200.00, "=E21*F21"]
    ]
    
    for row_idx, r in enumerate(items, start=5):
        ws.row_dimensions[row_idx].height = 19
        for col_idx, val in enumerate(r, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = font_body
            cell.border = border_cell
            
            if col_idx in [1, 2, 4]:
                cell.alignment = align_center
            elif col_idx == 3:
                cell.alignment = align_left
            elif col_idx == 5:
                cell.alignment = align_right
                cell.number_format = "#,##0"
            elif col_idx in [6, 7]:
                cell.alignment = align_right
                cell.number_format = "S/. #,##0.00"
                
            if row_idx % 2 == 0:
                cell.fill = fill_zebra
                
    # Summary math rows
    subtot_row = 22
    ws.row_dimensions[subtot_row].height = 20
    ws.cell(row=subtot_row, column=3, value="SUBTOTAL").font = font_total
    ws.cell(row=subtot_row, column=7, value="=SUM(G5:G21)").font = font_total
    ws.cell(row=subtot_row, column=7).number_format = "S/. #,##0.00"
    ws.cell(row=subtot_row, column=7).alignment = align_right
    ws.cell(row=subtot_row, column=7).border = Border(top=thin_side, bottom=thin_side)
    
    igv_row = 23
    ws.row_dimensions[igv_row].height = 20
    ws.cell(row=igv_row, column=3, value="IGV (18%)").font = font_total
    ws.cell(row=igv_row, column=7, value="=G22*0.18").font = font_total
    ws.cell(row=igv_row, column=7).number_format = "S/. #,##0.00"
    ws.cell(row=igv_row, column=7).alignment = align_right
    ws.cell(row=igv_row, column=7).border = Border(top=thin_side, bottom=thin_side)
    
    total_row = 24
    ws.row_dimensions[total_row].height = 22
    ws.cell(row=total_row, column=3, value="TOTAL GENERAL PRESUPUESTADO (S/.)").font = font_total
    ws.cell(row=total_row, column=7, value="=G22+G23").font = font_total
    ws.cell(row=total_row, column=7).number_format = "S/. #,##0.00"
    ws.cell(row=total_row, column=7).alignment = align_right
    ws.cell(row=total_row, column=7).border = border_total
    ws.cell(row=total_row, column=7).fill = fill_accent
    
    wb.save(os.path.join(base_dir, "04_metrados_y_presupuesto", "presupuesto.xlsx"))
    print("presupuesto.xlsx generado.")

if __name__ == "__main__":
    print("Iniciando generación de hojas de cálculo...")
    generate_cuadro_cargas()
    generate_maxima_demanda()
    generate_metrados()
    generate_presupuesto()
    print("¡Hojas de cálculo generadas con éxito!")
