#!/usr/bin/env python3
"""Cotizador automático multi-proveedor para materiales eléctricos."""

from __future__ import annotations

import argparse
import csv
import hashlib
import html
import json
import math
import os
import re
import shutil
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

BASE_DIR = Path(__file__).resolve().parent
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))

from modelos import ComparativaItem, MaterialBOM, ResultadoProveedor
from normalizador_materiales import (
    detectar_categoria,
    extraer_especificaciones,
    normalizar_nombre_material,
    score_tecnico,
)
from proveedores import PROVEEDORES_DISPONIBLES
from conversor_unidades import (
    detectar_unidad_comercial,
    extraer_contenido_comercial,
    calcular_compra_comercial,
    marcar_confianza_conversion,
    normalizar_unidad,
)

CACHE_PATH = BASE_DIR / ".cache" / "precios_multi_proveedor.json"
CACHE_CONTRACT_VERSION = 2
DEFAULT_PROVEEDORES = ["promart", "sodimac", "maestro", "mercadolibre"]


def cargar_json(path: Path) -> Any:
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def guardar_json(path: Path, data: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


def bom_hash(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(65536), b""):
            h.update(chunk)
    return h.hexdigest()[:16]


def cargar_bom(path: Path, max_materiales: Optional[int] = None) -> Tuple[Dict[str, Any], List[MaterialBOM]]:
    raw = cargar_json(path)
    if isinstance(raw, dict):
        materiales_raw = raw.get("materiales") or raw.get("items") or []
    elif isinstance(raw, list):
        materiales_raw = raw
    else:
        raise ValueError("Formato de BOM no reconocido")

    materiales: List[MaterialBOM] = []
    for idx, item in enumerate(materiales_raw, 1):
        if max_materiales and len(materiales) >= max_materiales:
            break
        if not isinstance(item, dict):
            continue
        nombre = (
            item.get("item")
            or item.get("descripcion")
            or item.get("nombre")
            or item.get("material")
            or ""
        )
        if not nombre:
            continue
        normalizado = normalizar_nombre_material(nombre)
        categoria = item.get("categoria") or detectar_categoria(nombre)
        cantidad = item.get("cantidad", 0) or 0
        try:
            cantidad = float(cantidad)
        except Exception:
            cantidad = 0.0
        material = MaterialBOM(
            id=item.get("codigo") or item.get("id") or f"M{idx:03d}",
            nombre_original=nombre,
            nombre_normalizado=normalizado,
            categoria=categoria,
            unidad=item.get("unidad", "und"),
            cantidad=cantidad,
            descripcion_uso=item.get("uso") or item.get("especificacion") or item.get("descripcion_uso") or "",
            especificaciones=extraer_especificaciones(nombre),
            precio_referencial=item.get("precio_unit_soles") or item.get("precio_unitario"),
        )
        materiales.append(material)
    return raw if isinstance(raw, dict) else {"materiales": raw}, materiales


def cargar_cache(path: Path = CACHE_PATH) -> Dict[str, Any]:
    if not path.exists():
        return {"version": 1, "entradas": {}}
    try:
        data = cargar_json(path)
        if "entradas" not in data:
            return {"version": 1, "entradas": {}}
        return data
    except Exception:
        return {"version": 1, "entradas": {}}


def guardar_cache(cache: Dict[str, Any], path: Path = CACHE_PATH) -> None:
    guardar_json(path, cache)


def cache_key(proveedor_slug: str, consulta: str) -> str:
    return f"{proveedor_slug}::{consulta.strip().lower()}"


def resultados_desde_cache(cache: Dict[str, Any], proveedor_slug: str, consulta: str) -> Optional[List[ResultadoProveedor]]:
    entrada = cache.get("entradas", {}).get(cache_key(proveedor_slug, consulta))
    if not entrada:
        return None
    resultados_raw = entrada.get("resultados", [])
    tiene_precio = any(r.get("precio") is not None for r in resultados_raw)
    if not tiene_precio and entrada.get("contract_version") != CACHE_CONTRACT_VERSION:
        return None
    resultados = []
    for raw in resultados_raw:
        r = ResultadoProveedor.from_dict(raw)
        r.metodo_extraccion = "cache"
        r.proveedor_slug = r.proveedor_slug or proveedor_slug
        if not r.observaciones:
            r.observaciones = "Resultado reutilizado desde cache local"
        resultados.append(r)
    return resultados


def guardar_resultados_cache(
    cache: Dict[str, Any],
    proveedor_slug: str,
    consulta: str,
    resultados: List[ResultadoProveedor],
    hash_bom: str,
) -> None:
    cache.setdefault("entradas", {})[cache_key(proveedor_slug, consulta)] = {
        "consulta": consulta,
        "proveedor": proveedor_slug,
        "fecha": datetime.now().isoformat(timespec="seconds"),
        "bom_hash": hash_bom,
        "contract_version": CACHE_CONTRACT_VERSION,
        "resultados": [r.to_dict() for r in resultados],
    }


def cargar_reglas() -> Dict[str, Any]:
    path = BASE_DIR / "data" / "reglas_matching.json"
    if not path.exists():
        return {
            "pesos": {"precio": 0.5, "tecnico": 0.3, "proveedor": 0.1, "disponibilidad": 0.1},
            "rangos_precio_referenciales": {},
        }
    return cargar_json(path)


def cargar_equivalencias_superiores() -> Dict[str, Any]:
    path = BASE_DIR / "data" / "equivalencias_superiores.json"
    if not path.exists():
        return {}
    try:
        return cargar_json(path)
    except Exception:
        return {}


def cargar_precios_manuales() -> List[Dict[str, Any]]:
    path = BASE_DIR / "data" / "precios_manual_verificado.json"
    if not path.exists():
        return []
    try:
        data = cargar_json(path)
        if isinstance(data, list):
            return [x for x in data if isinstance(x, dict)]
        if isinstance(data, dict) and data.get("material_id"):
            return [data]
        if isinstance(data, dict):
            return [x for x in data.get("precios", []) if isinstance(x, dict)]
        return []
    except Exception:
        return []


def disponibilidad_score(valor: Optional[str]) -> float:
    if not valor:
        return 0.65
    v = str(valor).lower()
    if any(x in v for x in ["disponible", "instock", "in stock", "stock"]):
        return 1.0
    if any(x in v for x in ["agotado", "outofstock", "sin stock", "no disponible"]):
        return 0.1
    if "no extraido" in v:
        return 0.25
    return 0.65


def precio_sospechoso(precio: Optional[float], categoria: str, reglas: Dict[str, Any]) -> bool:
    if precio is None:
        return False
    rangos = reglas.get("rangos_precio_referenciales", {})
    low_high = rangos.get(categoria)
    if not low_high:
        return False
    low, high = low_high
    return precio < low or precio > high


def unidad_comparable(material: MaterialBOM, resultado: ResultadoProveedor) -> Tuple[bool, str]:
    """Evita comparar precios por rollo/tubo contra cantidades en metros si no se ha convertido."""
    if getattr(resultado, "confianza_conversion", 0.0) >= 0.5:
        return True, ""
    unidad = (material.unidad or "").strip().lower()
    producto = f"{resultado.producto or ''} {resultado.unidad_detectada or ''}".lower()
    if unidad in {"m", "mt", "metro", "metros"}:
        if (resultado.unidad_detectada or "").strip().lower() in {"m", "mt", "metro", "metros"}:
            return True, ""
        if any(x in producto for x in ["por metro", "/m", " x metro", "1 m ", "1m "]):
            return True, ""
        if any(x in producto for x in [
            "rollo", "bobina", "100m", "100 m", "90m", "90 m", "50m", "50 m",
            "tubo", "barra", "varilla", "unidad", " un",
        ]):
            return False, "unidad no comparable con cantidad en metros"
        return False, "unidad de venta no visible para precio por metro"
    return True, ""


def _superior_inmediato(categoria: str, valor: float) -> Optional[float]:
    regla = cargar_equivalencias_superiores().get("reglas", {}).get(categoria, {})
    secuencia = [float(x) for x in regla.get("secuencia", [])]
    superiores = [x for x in secuencia if x > float(valor)]
    return min(superiores) if superiores else None


def clasificar_coincidencia_tecnica(material: MaterialBOM, resultado: ResultadoProveedor) -> None:
    if not resultado.producto:
        return
    objetivo = material.especificaciones
    encontrado = extraer_especificaciones(resultado.producto)
    categoria = material.categoria
    categoria_producto = detectar_categoria(resultado.producto)
    if categoria_producto != categoria and categoria != "accesorios":
        resultado.tipo_coincidencia = "similar_pendiente_revision"
        resultado.estado_proveedor = "baja_confianza"
        resultado.motivo_fallo = f"Categoria no comparable: {categoria_producto}."
        return

    if categoria == "cables" and objetivo.get("tipo_conductor"):
        titulo_limpio = (resultado.producto or "").lower()
        if not re.search(r"\b(?:thw|tw)\b", titulo_limpio):
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
            resultado.motivo_fallo = "El producto no confirma aislamiento THW/TW."
            return

    if categoria in {"cables", "ductos/tuberias", "tableros electricos"}:
        parametro = {
            "cables": "seccion_mm2",
            "ductos/tuberias": "diametro_mm",
            "tableros electricos": "capacidad_polos",
        }[categoria]
        valor_objetivo = objetivo.get(parametro)
        valor_encontrado = encontrado.get(parametro)
        if valor_objetivo is not None and valor_encontrado is not None:
            if float(valor_encontrado) == float(valor_objetivo):
                return
            inmediato = _superior_inmediato(categoria, float(valor_objetivo))
            if inmediato is not None and float(valor_encontrado) == inmediato:
                resultado.tipo_coincidencia = "equivalente_superior"
                resultado.observaciones = (
                    f"{resultado.observaciones}; superior inmediato {valor_objetivo:g} -> {inmediato:g}, revisar"
                ).strip("; ")
                return
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
        elif resultado.tipo_coincidencia == "equivalente_superior":
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
            resultado.motivo_fallo = "No se pudo verificar la medida del supuesto superior inmediato."

    if categoria == "interruptores termomagneticos":
        for parametro in ("amperaje_a", "polos"):
            if objetivo.get(parametro) is not None and encontrado.get(parametro) != objetivo.get(parametro):
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = "ITM con amperaje o numero de polos distinto al BOM."
                return

    if categoria == "interruptores diferenciales":
        for parametro in ("sensibilidad_ma", "polos", "amperaje_a"):
            if objetivo.get(parametro) is not None and encontrado.get(parametro) != objetivo.get(parametro):
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = "Diferencial con corriente, sensibilidad o polos distintos; requiere revision."
                return

    if categoria == "cajas electricas":
        nombre_caja = material.nombre_original.lower()
        producto_caja = (resultado.producto or "").lower()
        for forma in ("octogonal", "rectangular", "estanca"):
            if forma in nombre_caja and forma not in producto_caja:
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = f"La caja encontrada no confirma formato {forma}."
                return
        if "registro" in nombre_caja and "tierra" in nombre_caja:
            if "registro" not in producto_caja or not any(
                palabra in producto_caja for palabra in ("tierra", "pozo", "spat")
            ):
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = "La caja no confirma uso como registro de puesta a tierra."
                return

    if categoria in {"tomacorrientes", "interruptores simples/dobles/conmutados"}:
        amperaje_objetivo = objetivo.get("amperaje_a")
        amperaje_encontrado = encontrado.get("amperaje_a")
        if amperaje_objetivo is not None and amperaje_encontrado is not None:
            if float(amperaje_objetivo) != float(amperaje_encontrado):
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = "El amperaje del aparato no coincide con el metrado."
                return

    if categoria == "accesorios" and objetivo.get("diametro_mm") is not None:
        if encontrado.get("diametro_mm") != objetivo.get("diametro_mm"):
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
            resultado.motivo_fallo = "El diametro del accesorio no coincide con la tuberia."
            return
    if categoria == "accesorios":
        nombre_accesorio = material.nombre_original.lower()
        producto_accesorio = (resultado.producto or "").lower()
        for tipo_accesorio in ("union", "curva"):
            if tipo_accesorio in nombre_accesorio and tipo_accesorio not in producto_accesorio:
                resultado.tipo_coincidencia = "similar_pendiente_revision"
                resultado.estado_proveedor = "baja_confianza"
                resultado.motivo_fallo = f"El producto no confirma ser {tipo_accesorio} para tuberia."
                return
        if "pvc sap" in nombre_accesorio and not any(
            termino in producto_accesorio for termino in ("sap", "electr", "tubo")
        ):
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
            resultado.motivo_fallo = "El accesorio no confirma compatibilidad con tuberia PVC SAP electrica."
            return

    nombre_material_tierra = material.nombre_original.lower()
    es_electrodo = (
        "varilla" in nombre_material_tierra
        and "conector" not in nombre_material_tierra
        and "abrazadera" not in nombre_material_tierra
    )
    if categoria == "puesta a tierra" and es_electrodo:
        producto_limpio = (resultado.producto or "").lower()
        identidad_valida = (
            "varilla" in producto_limpio
            and any(
                termino in producto_limpio
                for termino in ("puesta a tierra", "electrodo", "jabalina", "copperweld", "cobre")
            )
        )
        longitud_bom = re.search(r"x\s*(\d+(?:[.,]\d+)?)\s*m\b", material.nombre_original.lower())
        longitud_producto = re.search(r"x\s*(\d+(?:[.,]\d+)?)\s*m\b", producto_limpio)
        longitud_valida = bool(
            longitud_bom
            and longitud_producto
            and float(longitud_bom.group(1).replace(",", "."))
            == float(longitud_producto.group(1).replace(",", "."))
        )
        if not identidad_valida or not longitud_valida:
            resultado.tipo_coincidencia = "similar_pendiente_revision"
            resultado.estado_proveedor = "baja_confianza"
            resultado.motivo_fallo = (
                "La varilla no confirma ser electrodo de puesta a tierra de la longitud requerida."
            )
            return


def puntuar_resultados(
    material: MaterialBOM,
    resultados: List[ResultadoProveedor],
    reglas: Dict[str, Any],
) -> None:
    # La compra real manda: incluye redondeo a rollos, tubos o paquetes completos.
    precios = []
    for r in resultados:
        p = r.subtotal_compra_real if r.subtotal_compra_real is not None else r.precio
        if p and p > 0:
            precios.append(p)
            
    min_precio = min(precios) if precios else None
    pesos = reglas.get("pesos", {})
    wp = float(pesos.get("precio", 0.50))
    wt = float(pesos.get("tecnico", 0.30))
    wprov = float(pesos.get("proveedor", 0.10))
    wdisp = float(pesos.get("disponibilidad", 0.10))

    for r in resultados:
        producto_texto = " ".join(x for x in [r.producto or "", r.marca or ""] if x)
        tecnico, notas = score_tecnico(material.nombre_normalizado, material.especificaciones, producto_texto)
        if not r.producto:
            tecnico = 0.0
        elif producto_texto:
            categoria_producto = detectar_categoria(producto_texto)
            if categoria_producto != material.categoria and material.categoria != "accesorios":
                tecnico *= 0.35
                notas = (
                    (notas + "; " if notas else "")
                    + f"categoria dudosa: {categoria_producto}"
                )
        if r.tipo_coincidencia == "equivalente_superior" and detectar_categoria(producto_texto) == material.categoria:
            tecnico = max(tecnico, 0.68)
            notas = (notas + "; " if notas else "") + "superior inmediato permitido por regla"
        unidad_ok, unidad_nota = unidad_comparable(material, r)
        if not unidad_ok and r.precio is not None:
            tecnico *= 0.20
            notas = (notas + "; " if notas else "") + unidad_nota
            r.confianza = min(r.confianza, 0.35)
        sospechoso = precio_sospechoso(r.precio_equivalente_unitario if r.precio_equivalente_unitario is not None else r.precio, material.categoria, reglas)
        if sospechoso:
            tecnico *= 0.75
            notas = (notas + "; " if notas else "") + "precio fuera de rango referencial"
        r.score_tecnico = round(tecnico, 4)
        r.score_disponibilidad = round(disponibilidad_score(r.disponibilidad), 4)
        
        p_eval = r.subtotal_compra_real if r.subtotal_compra_real is not None else r.precio
        if p_eval and min_precio:
            r.score_precio = round(min(1.0, min_precio / p_eval), 4)
        else:
            r.score_precio = 0.0
            
        if sospechoso:
            r.score_precio = round(r.score_precio * 0.25, 4)
        if not unidad_ok:
            r.score_precio = round(r.score_precio * 0.20, 4)
        r.score_proveedor = r.score_proveedor or 0.5
        r.score_total = round(
            wp * r.score_precio
            + wt * r.score_tecnico
            + wprov * r.score_proveedor
            + wdisp * r.score_disponibilidad,
            4,
        )
        if notas and r.observaciones:
            r.observaciones = f"{r.observaciones}; {notas}"
        elif notas:
            r.observaciones = notas
        if r.precio is not None:
            if r.confianza_conversion <= 0:
                r.estado_proveedor = "unidad_no_convertible"
            elif r.tipo_coincidencia == "similar_pendiente_revision":
                r.estado_proveedor = "baja_confianza"
            elif r.score_tecnico < 0.45 or r.score_total < 0.45:
                r.estado_proveedor = "baja_confianza"
            else:
                r.estado_proveedor = "ok_precio"


def elegir_recomendado(material: MaterialBOM, resultados: List[ResultadoProveedor]) -> Tuple[Optional[ResultadoProveedor], str]:
    candidatos = []
    for r in resultados:
        if r.precio is None or not r.url:
            continue
        if r.score_tecnico < 0.45 or r.score_total < 0.45:
            continue
            
        tipo = getattr(r, "tipo_coincidencia", "exacta")
        if tipo in {"exacta", "comercial_convertida"}:
            if getattr(r, "confianza_conversion", 1.0) >= 0.5:
                candidatos.append(r)
        elif tipo == "equivalente_superior":
            if material.categoria in {"cables", "ductos/tuberias", "tableros electricos"}:
                if r.score_tecnico >= 0.50 and r.score_total >= 0.50 and getattr(r, "confianza_conversion", 1.0) >= 0.5:
                    candidatos.append(r)
                    
    if not candidatos:
        con_precio = [r for r in resultados if r.precio is not None and r.url]
        if con_precio:
            mejor = max(con_precio, key=lambda r: (r.score_tecnico, r.score_total))
            return None, (
                "Pendiente: hay precios, pero ninguno alcanza confianza tecnica minima. "
                f"Mejor candidato: {mejor.proveedor}, score tecnico {mejor.score_tecnico:.2f}, "
                f"producto: {mejor.producto}."
            )
        return None, "Pendiente: ningun proveedor devolvio precio verificable."
        
    recomendado = max(
        candidatos,
        key=lambda r: (
            r.score_total,
            r.score_tecnico,
            -float(r.subtotal_compra_real or float("inf")),
        ),
    )
    motivo = (
        f"Mejor balance precio/tecnica: score {recomendado.score_total:.2f}, "
        f"coincidencia tecnica {recomendado.score_tecnico:.2f}, "
        f"precio S/ {recomendado.precio:.2f} ({recomendado.unidad_comercial or 'und'})."
    )
    if recomendado.score_tecnico < 0.45:
        motivo += " Requiere revision manual por baja coincidencia tecnica."
    return recomendado, motivo


def cargar_fixture(proveedor_slug: str) -> Optional[str]:
    path = BASE_DIR / "tests" / "fixtures" / "html_muestras" / f"{proveedor_slug}.html"
    if path.exists():
        return path.read_text(encoding="utf-8")
    return None


def consultar_proveedor(
    proveedor_slug: str,
    proveedor: Any,
    material: MaterialBOM,
    args: argparse.Namespace,
    cache: Dict[str, Any],
    hash_bom: str,
    evidencias_dir: Path,
    evidencias_manifest: List[Dict[str, Any]],
    query_override: Optional[str] = None,
) -> List[ResultadoProveedor]:
    consulta = query_override or material.nombre_normalizado

    # Primero verificar si existe precio manual verificado
    manual_entries = cargar_precios_manuales()
    for entry in manual_entries:
        if (
            entry.get("material_id") == material.id
            and entry.get("proveedor") == proveedor_slug
            and entry.get("url")
            and float(entry.get("precio") or 0) > 0
        ):
            r = ResultadoProveedor(
                proveedor=proveedor.nombre,
                proveedor_slug=proveedor_slug,
                consulta=consulta,
                producto=entry.get("producto"),
                precio=entry.get("precio"),
                moneda=entry.get("moneda", "PEN"),
                unidad_detectada=entry.get("unidad_comercial"),
                url=entry.get("url"),
                disponibilidad="Disponible",
                fecha_consulta=entry.get("fecha", ""),
                metodo_extraccion="manual",
                confianza=1.0,
                observaciones=entry.get("observacion", "Precio manual verificado"),
                estado_proveedor="ok_precio",
                url_solicitada=entry.get("url"),
                url_final=entry.get("url"),
                tipo_enlace="producto",
            )
            # Aplicar la conversion comercial
            r.unidad_bom = material.unidad
            r.cantidad_bom = material.cantidad
            r.unidad_comercial = entry.get("unidad_comercial")
            r.contenido_comercial = float(entry.get("contenido_comercial", 1.0))
            r.unidad_contenido = entry.get("unidad_contenido")
            
            cant_com, subtotal_real, sobrante, precio_equiv = calcular_compra_comercial(
                cantidad_bom=material.cantidad,
                unidad_bom=material.unidad,
                precio_producto=r.precio,
                unidad_comercial=r.unidad_comercial,
                contenido_comercial=r.contenido_comercial
            )
            r.cantidad_comercial_a_comprar = cant_com
            r.precio_equivalente_unitario = precio_equiv
            r.subtotal_compra_real = subtotal_real
            r.sobrante = sobrante
            r.confianza_conversion = 1.0
            r.observacion_conversion = "Precio manual verificado"
            r.tipo_coincidencia = "exacta"
            
            registrar_evidencia_json(
                evidencias_dir,
                evidencias_manifest,
                proveedor_slug,
                material.id,
                consulta,
                "manual",
                entry,
            )
            return [r]

    if args.usar_cache and not args.refrescar_cache and not args.offline:
        cached = resultados_desde_cache(cache, proveedor_slug, consulta)
        if cached:
            registrar_evidencia_json(
                evidencias_dir,
                evidencias_manifest,
                proveedor_slug,
                material.id,
                consulta,
                "cache",
                {"cache_key": cache_key(proveedor_slug, consulta), "resultados": [r.to_dict() for r in cached]},
            )
            return cached[: args.max_resultados]

    if args.offline:
        if not args.usar_fixtures:
            return [proveedor.resultado_fallback(
                consulta,
                proveedor.build_search_url(consulta),
                "Modo offline sin fixtures; no se consulto internet",
                metodo="fallback",
            )]
        html_fixture = cargar_fixture(proveedor_slug)
        if not html_fixture:
            return [proveedor.resultado_fallback(
                consulta,
                proveedor.build_search_url(consulta),
                "Modo offline: fixture no disponible",
                metodo="fallback",
            )]
        proveedor.last_url = f"fixture://{proveedor_slug}.html"
        proveedor.last_html = html_fixture
        resultados = proveedor.parsear_resultados(html_fixture, consulta)[: args.max_resultados]
        registrar_evidencia_texto(
            evidencias_dir,
            evidencias_manifest,
            proveedor_slug,
            material.id,
            consulta,
            "fixture",
            html_fixture,
            proveedor.last_url,
        )
        if resultados:
            return resultados
        return [proveedor.resultado_fallback(
            consulta,
            proveedor.last_url,
            "Fixture sin resultados parseables",
            metodo="fallback",
        )]

    resultados = proveedor.buscar(consulta, args.max_resultados)
    if proveedor.last_html:
        registrar_evidencia_texto(
            evidencias_dir,
            evidencias_manifest,
            proveedor_slug,
            material.id,
            consulta,
            "html",
            proveedor.last_html,
            proveedor.last_url or proveedor.build_search_url(consulta),
            status=proveedor.last_status,
        )
    else:
        registrar_evidencia_json(
            evidencias_dir,
            evidencias_manifest,
            proveedor_slug,
            material.id,
            consulta,
            "error",
            {
                "url": proveedor.last_url or proveedor.build_search_url(consulta),
                "error": proveedor.last_error,
                "resultados": [r.to_dict() for r in resultados],
            },
        )
    for intento in getattr(proveedor, "request_history", []):
        evidencias_manifest.append({
            "material_id": material.id,
            "material": material.nombre_original,
            "proveedor": proveedor_slug,
            "consulta": consulta,
            "tipo": "diagnostico_http",
            **intento,
            "fecha": datetime.now().isoformat(timespec="seconds"),
        })
    for resultado in resultados:
        resultado.proveedor_slug = resultado.proveedor_slug or proveedor_slug
        resultado.resultados_crudos = len([r for r in resultados if r.producto or r.precio is not None])
    guardar_resultados_cache(cache, proveedor_slug, consulta, resultados, hash_bom)
    return resultados[: args.max_resultados]


def nombre_evidencia(proveedor_slug: str, material_id: str, tipo: str, consulta: str = "") -> str:
    safe_id = "".join(c if c.isalnum() or c in "-_" else "_" for c in material_id)
    digest = hashlib.sha1(consulta.encode("utf-8")).hexdigest()[:8] if consulta else "consulta"
    return f"{safe_id}_{proveedor_slug}_{digest}.{tipo}"


def registrar_evidencia_texto(
    evidencias_dir: Path,
    manifest: List[Dict[str, Any]],
    proveedor_slug: str,
    material_id: str,
    consulta: str,
    tipo: str,
    contenido: str,
    url: str,
    status: Optional[int] = None,
) -> None:
    raw_dir = evidencias_dir / "raw"
    raw_dir.mkdir(parents=True, exist_ok=True)
    filename = nombre_evidencia(
        proveedor_slug,
        material_id,
        "html" if tipo in {"html", "fixture"} else "txt",
        consulta,
    )
    path = raw_dir / filename
    path.write_text(contenido[:700000], encoding="utf-8", errors="replace")
    manifest.append({
        "material_id": material_id,
        "proveedor": proveedor_slug,
        "consulta": consulta,
        "tipo": tipo,
        "url": url,
        "status": status,
        "archivo": str(path),
        "fecha": datetime.now().isoformat(timespec="seconds"),
    })


def registrar_evidencia_json(
    evidencias_dir: Path,
    manifest: List[Dict[str, Any]],
    proveedor_slug: str,
    material_id: str,
    consulta: str,
    tipo: str,
    data: Dict[str, Any],
) -> None:
    evidencias_dir.mkdir(parents=True, exist_ok=True)
    filename = nombre_evidencia(proveedor_slug, material_id, "json", consulta)
    path = evidencias_dir / filename
    guardar_json(path, data)
    manifest.append({
        "material_id": material_id,
        "proveedor": proveedor_slug,
        "consulta": consulta,
        "tipo": tipo,
        "url": data.get("url") or "",
        "archivo": str(path),
        "fecha": datetime.now().isoformat(timespec="seconds"),
    })


def obtener_consultas_fase2(material: MaterialBOM) -> List[str]:
    categoria = material.categoria
    specs = material.especificaciones
    queries = []
    
    from normalizador_materiales import fmt_num
    
    if categoria == "cables":
        seccion = specs.get("seccion_mm2")
        if seccion:
            s_str = fmt_num(seccion)
            queries.append(f"rollo cable thw {s_str} mm2")
            queries.append(f"cable thw {s_str} mm2 100m")
            queries.append(f"cable electrico {s_str} mm2 rollo")
            queries.append(f"conductor cobre {s_str} mm2")
    elif categoria == "ductos/tuberias":
        diam = specs.get("diametro_mm")
        if diam:
            d_str = fmt_num(diam)
            queries.append(f"tubo pvc sap {d_str} mm")
            queries.append(f"tubo pvc electrico {d_str} mm")
            queries.append(f"tubo conduit pvc {d_str} mm")
            queries.append(f"tubo electrico {d_str} mm 3 metros")
    elif categoria == "interruptores termomagneticos":
        amp = specs.get("amperaje_a")
        polos = specs.get("polos")
        if amp and polos:
            a_str = fmt_num(amp)
            queries.extend([
                f"interruptor termomagnetico {polos}p {a_str}a",
                f"llave termomagnetica {polos}x{a_str}a",
                f"breaker {polos} polos {a_str}a",
                f"termomagnetico {a_str}a {polos} polos",
            ])
    elif categoria == "interruptores diferenciales":
        amp = specs.get("amperaje_a")
        polos = specs.get("polos")
        sensibilidad = specs.get("sensibilidad_ma")
        if amp and polos and sensibilidad:
            queries.extend([
                f"interruptor diferencial {polos}p {fmt_num(amp)}a {fmt_num(sensibilidad)}ma",
                f"diferencial {fmt_num(amp)}a {fmt_num(sensibilidad)}ma {polos} polos",
                f"disyuntor diferencial {polos}x{fmt_num(amp)}a {fmt_num(sensibilidad)}ma",
            ])
    elif categoria == "tableros electricos":
        capacidad = specs.get("capacidad_polos") or specs.get("polos")
        if capacidad:
            queries.extend([
                f"tablero electrico {capacidad} polos",
                f"tablero empotrable {capacidad} modulos",
                f"gabinete electrico {capacidad} circuitos",
            ])
    elif categoria == "puesta a tierra":
        n_lower = material.nombre_normalizado.lower()
        if "varilla" in n_lower or "electrodo" in n_lower:
            queries.append("varilla puesta a tierra 5/8")
            queries.append("electrodo puesta a tierra")
        elif "conductor" in n_lower or "cable" in n_lower:
            queries.append("cable cobre de tierra")
            queries.append("rollo cable tierra verde")
    elif categoria == "cajas electricas":
        n_lower = material.nombre_normalizado.lower()
        if "estanca" in n_lower:
            queries.extend(["caja estanca electrica ip55", "caja de paso estanca ip55"])
        elif "registro" in n_lower and "tierra" in n_lower:
            queries.extend(["caja registro puesta a tierra", "caja inspeccion pozo a tierra"])
        elif "rectangular" in n_lower:
            queries.append("caja rectangular pvc")
            queries.append("paquete caja rectangular")
        elif "octogonal" in n_lower:
            queries.append("caja octogonal pvc")
            queries.append("paquete caja octogonal")
        if "estanca" not in n_lower and "registro" not in n_lower:
            queries.append("caja electrica pvc")
    elif categoria == "accesorios":
        n_lower = material.nombre_normalizado.lower()
        if "cinta" in n_lower:
            queries.append("rollo cinta aislante")
            queries.append("cinta aislante 3m")
        elif "union" in n_lower:
            diam = specs.get("diametro_mm")
            if diam:
                queries.extend([
                    f"union pvc sap {fmt_num(diam)} mm",
                    f"union para tubo electrico {fmt_num(diam)} mm",
                ])
        elif "curva" in n_lower:
            diam = specs.get("diametro_mm")
            if diam:
                queries.extend([
                    f"curva pvc sap {fmt_num(diam)} mm",
                    f"curva para tubo electrico {fmt_num(diam)} mm",
                ])
    consultas_unicas = []
    vistos = {material.nombre_normalizado.strip().lower()}
    for query in queries:
        clave = query.strip().lower()
        if clave and clave not in vistos:
            consultas_unicas.append(query)
            vistos.add(clave)
    return consultas_unicas


def obtener_consulta_fase3(material: MaterialBOM) -> Optional[Tuple[str, str, str]]:
    categoria = material.categoria
    specs = material.especificaciones
    
    eq_rules = cargar_equivalencias_superiores().get("reglas", {})
    rule = eq_rules.get(categoria)
    if not rule:
        return None
        
    tipo = rule.get("tipo")
    from normalizador_materiales import fmt_num
    
    if tipo == "secuencia":
        param = rule.get("parametro")
        val = specs.get(param)
        if val is not None:
            seq = rule.get("secuencia", [])
            try:
                val_f = float(val)
                bigger = [x for x in seq if x > val_f]
                if bigger:
                    next_val = min(bigger)
                    val_str = fmt_num(val_f)
                    next_str = fmt_num(next_val)
                    
                    # Reemplazo en el nombre
                    query = re.sub(r"\b" + re.escape(val_str) + r"\b", next_str, material.nombre_normalizado)
                    
                    obs = rule.get("observacion", "Seccion superior sugerida.").replace("{val}", val_str).replace("{next_val}", next_str)
                    tipo_coinc = rule.get("tipo_coincidencia", "equivalente_superior")
                    return query, obs, tipo_coinc
            except Exception:
                pass
                
    elif tipo == "diferencial_flex":
        sens = specs.get("sensibilidad_ma")
        amp = specs.get("amperaje_a")
        if sens == 30 and amp is not None:
            if amp == 25:
                query = material.nombre_normalizado.replace("25a", "40a")
                obs = rule.get("observacion", "Diferencial de amperaje superior.")
                tipo_coinc = rule.get("tipo_coincidencia", "similar_pendiente_revision")
                return query, obs, tipo_coinc
                
    return None


def aplicar_conversion_comercial_a_resultado(r: ResultadoProveedor, material: MaterialBOM) -> None:
    if r.precio is None:
        r.unidad_bom = material.unidad
        r.cantidad_bom = material.cantidad
        return
        
    # Detectar unidad comercial
    unidad_comercial = detectar_unidad_comercial(r.producto or "", material.categoria)
    
    # Extraer contenido y confianza
    contenido, und_contenido, confianza_conv, obs_conv = extraer_contenido_comercial(r.producto or "", material.categoria)
    if material.categoria == "cables" and contenido > 1 and unidad_comercial == "m":
        unidad_comercial = "rollo"
        obs_conv = f"{obs_conv} Unidad comercial inferida como rollo por longitud total del producto."
    
    # Calcular compra comercial
    cant_com, subtotal_real, sobrante, precio_equiv = calcular_compra_comercial(
        cantidad_bom=material.cantidad,
        unidad_bom=material.unidad,
        precio_producto=r.precio,
        unidad_comercial=unidad_comercial,
        contenido_comercial=contenido
    )
    
    r.unidad_bom = material.unidad
    r.cantidad_bom = material.cantidad
    r.unidad_comercial = unidad_comercial
    r.contenido_comercial = contenido
    r.unidad_contenido = und_contenido
    r.cantidad_comercial_a_comprar = cant_com
    r.precio_equivalente_unitario = precio_equiv
    r.subtotal_compra_real = subtotal_real
    r.sobrante = sobrante
    r.confianza_conversion = marcar_confianza_conversion(
        material.unidad,
        unidad_comercial,
        contenido,
        confianza_conv,
    )
    r.observacion_conversion = obs_conv
    if r.confianza_conversion <= 0:
        r.estado_proveedor = "unidad_no_convertible"
        r.motivo_fallo = obs_conv


def tipo_coincidencia_por_conversion(resultado: ResultadoProveedor) -> str:
    if resultado.confianza_conversion <= 0:
        return "similar_pendiente_revision"
    if (
        normalizar_unidad(resultado.unidad_bom) == normalizar_unidad(resultado.unidad_comercial)
        and resultado.contenido_comercial == 1.0
    ):
        return "exacta"
    return "comercial_convertida"


def construir_comparativa(
    materiales: List[MaterialBOM],
    proveedores_slugs: List[str],
    args: argparse.Namespace,
    cache: Dict[str, Any],
    hash_bom: str,
    output_dir: Path,
) -> Tuple[List[ComparativaItem], List[Dict[str, Any]]]:
    reglas = cargar_reglas()
    evidencias_dir = output_dir / "evidencias"
    manifest: List[Dict[str, Any]] = []
    comparativas: List[ComparativaItem] = []
    
    proveedores = {
        slug: PROVEEDORES_DISPONIBLES[slug](timeout=args.timeout, delay=args.delay)
        for slug in proveedores_slugs
    }

    for idx, material in enumerate(materiales, 1):
        print(f"[{idx}/{len(materiales)}] {material.nombre_original} -> {material.nombre_normalizado}")
        resultados_item: List[ResultadoProveedor] = []
        
        for slug, proveedor in proveedores.items():
            print(f"  - {proveedor.nombre}", flush=True)
            
            # FASE 1: Búsqueda exacta
            resultados = consultar_proveedor(
                slug, proveedor, material, args, cache, hash_bom, evidencias_dir, manifest, query_override=None
            )
            
            # Aplicar conversiones
            for r in resultados:
                aplicar_conversion_comercial_a_resultado(r, material)
                r.tipo_coincidencia = tipo_coincidencia_por_conversion(r)
                if r.confianza_conversion == 0.0:
                    r.tipo_coincidencia = "similar_pendiente_revision"
                clasificar_coincidencia_tecnica(material, r)
                    
            # Evaluar si Fase 1 tiene candidato aceptable
            puntuar_resultados(material, resultados, reglas)
            validas_f1 = [
                r for r in resultados
                if r.precio is not None
                and r.url
                and r.score_tecnico >= 0.45
                and r.score_total >= 0.45
                and r.confianza_conversion >= 0.5
                and r.tipo_coincidencia != "similar_pendiente_revision"
                and r.metodo_extraccion != "fallback"
            ]
            
            if validas_f1:
                resultados_item.extend(resultados)
                continue
            if any(r.estado_proveedor in {"bloqueado", "timeout"} for r in resultados):
                resultados_item.extend(resultados)
                continue
                
            # FASE 2: Búsqueda con equivalentes comerciales
            consultas_f2 = obtener_consultas_fase2(material)
            encontrado_f2 = False
            
            for q_f2 in consultas_f2:
                print(f"    * Probando Fase 2: {q_f2}", flush=True)
                res_f2 = consultar_proveedor(
                    slug, proveedor, material, args, cache, hash_bom, evidencias_dir, manifest, query_override=q_f2
                )
                for r in res_f2:
                    aplicar_conversion_comercial_a_resultado(r, material)
                    r.tipo_coincidencia = tipo_coincidencia_por_conversion(r)
                    if r.confianza_conversion == 0.0:
                        r.tipo_coincidencia = "similar_pendiente_revision"
                    clasificar_coincidencia_tecnica(material, r)
                        
                puntuar_resultados(material, res_f2, reglas)
                validas_f2 = [
                    r for r in res_f2
                    if r.precio is not None
                    and r.url
                    and r.score_tecnico >= 0.45
                    and r.score_total >= 0.45
                    and r.confianza_conversion >= 0.5
                    and r.tipo_coincidencia != "similar_pendiente_revision"
                    and r.metodo_extraccion != "fallback"
                ]
                
                if validas_f2:
                    resultados_item.extend(res_f2)
                    encontrado_f2 = True
                    break
                    
            if encontrado_f2:
                continue
                
            # FASE 3: Búsqueda con superior inmediato
            fase3_info = obtener_consulta_fase3(material)
            if fase3_info:
                q_f3, obs_f3, tipo_coinc_f3 = fase3_info
                print(f"    * Probando Fase 3: {q_f3}", flush=True)
                res_f3 = consultar_proveedor(
                    slug, proveedor, material, args, cache, hash_bom, evidencias_dir, manifest, query_override=q_f3
                )
                for r in res_f3:
                    aplicar_conversion_comercial_a_resultado(r, material)
                    r.tipo_coincidencia = "similar_pendiente_revision"
                    r.observaciones = f"{r.observaciones or ''}; {obs_f3}".strip("; ")
                    if r.confianza_conversion == 0.0:
                        r.tipo_coincidencia = "similar_pendiente_revision"
                    clasificar_coincidencia_tecnica(material, r)
                        
                puntuar_resultados(material, res_f3, reglas)
                resultados_item.extend(res_f3)
            else:
                # Si nada funcionó, agregamos los de Fase 1 (fallbacks)
                resultados_item.extend(resultados)

        # Repuntuar para que todos compitan justamente con precio equivalente
        puntuar_resultados(material, resultados_item, reglas)
        recomendado, motivo = elegir_recomendado(material, resultados_item)
        
        comparativas.append(ComparativaItem(
            material_bom=material,
            resultados_por_proveedor=resultados_item,
            resultado_recomendado=recomendado,
            motivo_recomendacion=motivo,
        ))
    return comparativas, manifest


def resultado_aceptable_tienda(r: ResultadoProveedor) -> bool:
    return bool(
        r.precio is not None
        and r.url
        and r.score_tecnico >= 0.45
        and r.score_total >= 0.45
        and r.confianza_conversion >= 0.5
        and r.tipo_coincidencia != "similar_pendiente_revision"
        and r.metodo_extraccion != "fallback"
    )


def clave_consolidacion(material: MaterialBOM) -> Tuple[str, str, str, str]:
    nombre = material.nombre_original.lower()
    uso_especial = "tierra" if "tierra" in nombre or "verde/amarillo" in nombre else "general"
    return (
        material.nombre_normalizado,
        material.categoria,
        normalizar_unidad(material.unidad),
        uso_especial,
    )


def construir_plan_compra_consolidado(
    comparativas: List[ComparativaItem],
    proveedores: Optional[set[str]] = None,
) -> List[Dict[str, Any]]:
    grupos: Dict[Tuple[str, str, str, str], List[ComparativaItem]] = defaultdict(list)
    for comp in comparativas:
        grupos[clave_consolidacion(comp.material_bom)].append(comp)

    plan: List[Dict[str, Any]] = []
    for grupo in grupos.values():
        cantidad_total = sum(comp.material_bom.cantidad for comp in grupo)
        candidatos: Dict[Tuple[Any, ...], ResultadoProveedor] = {}
        for comp in grupo:
            for resultado in comp.resultados_por_proveedor:
                if not resultado_aceptable_tienda(resultado):
                    continue
                if proveedores is not None and slug_resultado(resultado) not in proveedores:
                    continue
                clave = (
                    slug_resultado(resultado),
                    resultado.producto,
                    resultado.precio,
                    resultado.unidad_comercial,
                    resultado.contenido_comercial,
                )
                anterior = candidatos.get(clave)
                if anterior is None or resultado.score_total > anterior.score_total:
                    candidatos[clave] = resultado

        opciones: List[Tuple[float, ResultadoProveedor, float, float, float]] = []
        for resultado in candidatos.values():
            cantidad_comercial, subtotal, sobrante, precio_equivalente = calcular_compra_comercial(
                cantidad_total,
                grupo[0].material_bom.unidad,
                float(resultado.precio or 0.0),
                resultado.unidad_comercial or "unidad",
                float(resultado.contenido_comercial or 1.0),
            )
            opciones.append((subtotal, resultado, cantidad_comercial, sobrante, precio_equivalente))

        mejor = min(
            opciones,
            key=lambda opcion: (opcion[0], -opcion[1].score_tecnico, -opcion[1].score_total),
            default=None,
        )
        material_base = grupo[0].material_bom
        fila: Dict[str, Any] = {
            "material": material_base.nombre_normalizado,
            "descripcion_representativa": material_base.nombre_original,
            "material_ids": [comp.material_bom.id for comp in grupo],
            "partidas_bom": len(grupo),
            "cantidad_total_bom": cantidad_total,
            "unidad_bom": material_base.unidad,
            "categoria": material_base.categoria,
            "proveedor": None,
            "proveedor_slug": None,
            "producto": None,
            "precio_comercial": None,
            "unidad_comercial": None,
            "contenido_comercial": None,
            "unidad_contenido": None,
            "cantidad_comercial_a_comprar": None,
            "precio_equivalente_unitario": None,
            "subtotal_compra_real": None,
            "sobrante": None,
            "tipo_coincidencia": None,
            "url": None,
            "confianza": None,
            "observacion": "Sin oferta aceptable para el grupo consolidado.",
        }
        if mejor:
            subtotal, resultado, cantidad_comercial, sobrante, precio_equivalente = mejor
            fila.update({
                "proveedor": resultado.proveedor,
                "proveedor_slug": slug_resultado(resultado),
                "producto": resultado.producto,
                "precio_comercial": resultado.precio,
                "unidad_comercial": resultado.unidad_comercial,
                "contenido_comercial": resultado.contenido_comercial,
                "unidad_contenido": resultado.unidad_contenido,
                "cantidad_comercial_a_comprar": cantidad_comercial,
                "precio_equivalente_unitario": precio_equivalente,
                "subtotal_compra_real": subtotal,
                "sobrante": sobrante,
                "tipo_coincidencia": resultado.tipo_coincidencia,
                "url": resultado.url,
                "confianza": resultado.score_total,
                "observacion": resultado.observacion_conversion or resultado.observaciones,
            })
        plan.append(fila)
    return plan


def resumen_totales(comparativas: List[ComparativaItem]) -> Dict[str, Any]:
    total_materiales = len(comparativas)
    nombres_proveedores = ["Promart", "Sodimac", "Maestro", "Mercado Libre"]
    proveedores: Dict[str, Dict[str, Any]] = {
        nombre: {
            "total_acumulado": 0.0,
            "items_con_precio": 0,
            "items_sin_precio": 0,
            "materiales_pendientes": [],
            "estados": defaultdict(int),
        }
        for nombre in nombres_proveedores
    }
    plan_mixto = construir_plan_compra_consolidado(comparativas)
    recomendado_total = sum(
        float(fila["precio_equivalente_unitario"] or 0.0) * float(fila["cantidad_total_bom"])
        for fila in plan_mixto
        if fila["subtotal_compra_real"] is not None
    )
    total_compra_real_mixta = sum(
        float(fila["subtotal_compra_real"] or 0.0)
        for fila in plan_mixto
        if fila["subtotal_compra_real"] is not None
    )
    sin_precio: List[str] = []
    baja_confianza: List[Dict[str, Any]] = []
    encontrados_total = 0

    for comp in comparativas:
        mat = comp.material_bom
        por_proveedor: Dict[str, List[ResultadoProveedor]] = defaultdict(list)
        for resultado in comp.resultados_por_proveedor:
            por_proveedor[resultado.proveedor].append(resultado)

        for proveedor in nombres_proveedores:
            resultados = por_proveedor.get(proveedor, [])
            aceptados = [r for r in resultados if resultado_aceptable_tienda(r)]
            for resultado in resultados:
                proveedores[proveedor]["estados"][resultado.estado_proveedor] += 1
            if aceptados:
                mejor = max(aceptados, key=lambda r: (r.score_total, -float(r.subtotal_compra_real or 0)))
                proveedores[proveedor]["items_con_precio"] += 1
                mejor.resultados_aceptados = 1
            else:
                proveedores[proveedor]["items_sin_precio"] += 1
                proveedores[proveedor]["materiales_pendientes"].append(mat.nombre_original)

        recomendado = comp.resultado_recomendado
        if recomendado and recomendado.precio is not None:
            encontrados_total += 1
        else:
            sin_precio.append(mat.nombre_original)
            con_precio = [r for r in comp.resultados_por_proveedor if r.precio is not None]
            if con_precio:
                mejor = max(con_precio, key=lambda r: (r.score_tecnico, r.score_total))
                baja_confianza.append({
                    "material": mat.nombre_original,
                    "proveedor": mejor.proveedor,
                    "score_total": mejor.score_total,
                    "score_tecnico": mejor.score_tecnico,
                    "motivo": mejor.motivo_fallo or mejor.observaciones,
                })

    totales_por_proveedor: Dict[str, Dict[str, Any]] = {}
    proveedores_utiles = 0
    for proveedor, datos in proveedores.items():
        slug_proveedor = {
            "Promart": "promart",
            "Sodimac": "sodimac",
            "Maestro": "maestro",
            "Mercado Libre": "mercadolibre",
        }[proveedor]
        plan_proveedor = construir_plan_compra_consolidado(comparativas, {slug_proveedor})
        datos["total_acumulado"] = sum(
            float(fila["subtotal_compra_real"] or 0.0)
            for fila in plan_proveedor
            if fila["subtotal_compra_real"] is not None
        )
        cobertura_pct = round(
            datos["items_con_precio"] / total_materiales * 100.0,
            2,
        ) if total_materiales else 0.0
        if cobertura_pct >= 30.0:
            proveedores_utiles += 1
        if datos["items_con_precio"] == 0:
            estado = "sin cobertura"
            advertencia = "ADVERTENCIA: este proveedor no tuvo cobertura util. La comparativa no es completa."
            total = None
        elif cobertura_pct < 100.0:
            estado = "parcial"
            advertencia = "No comparar el total como presupuesto completo: faltan materiales."
            total = round(datos["total_acumulado"], 2)
        else:
            estado = "completo"
            advertencia = ""
            total = round(datos["total_acumulado"], 2)
        totales_por_proveedor[proveedor] = {
            "total": total,
            "total_parcial_encontrado": total,
            "items_con_precio": datos["items_con_precio"],
            "items_sin_precio": datos["items_sin_precio"],
            "cobertura_pct": cobertura_pct,
            "estado": estado,
            "advertencia": advertencia,
            "materiales_pendientes": datos["materiales_pendientes"],
            "estados_busqueda": dict(datos["estados"]),
        }

    comparativa_completa = proveedores_utiles >= 2
    totales_validos = [
        datos["total"]
        for datos in totales_por_proveedor.values()
        if datos["total"] is not None and datos["cobertura_pct"] == 100.0
    ]
    mas_caro = max(totales_validos) if totales_validos else None
    diferencia_pct = (
        (mas_caro - total_compra_real_mixta) / mas_caro * 100.0
        if mas_caro and total_compra_real_mixta
        else None
    )

    return {
        "total_recomendado_materiales": round(recomendado_total, 2),
        "total_compra_real_mixta": round(total_compra_real_mixta, 2),
        "cobertura_bom": f"{encontrados_total}/{total_materiales}",
        "cobertura_bom_pct": round(encontrados_total / total_materiales * 100.0, 2) if total_materiales else 0.0,
        "comparativa_completa": comparativa_completa,
        "proveedores_con_cobertura_util": proveedores_utiles,
        "mensaje_calidad": (
            "Comparativa con cobertura util en al menos dos proveedores."
            if comparativa_completa
            else f"Cotizacion parcial: solo se obtuvo cobertura util en {proveedores_utiles} proveedores."
        ),
        "totales_por_proveedor": totales_por_proveedor,
        "diferencia_porcentual_mas_caro_vs_recomendado": (
            round(diferencia_pct, 2) if diferencia_pct is not None else None
        ),
        "materiales_sin_precio": sin_precio,
        "materiales_baja_confianza": baja_confianza,
        "plan_compra_consolidado": plan_mixto,
    }


def generar_salidas(
    raw_bom: Dict[str, Any],
    materiales: List[MaterialBOM],
    comparativas: List[ComparativaItem],
    manifest: List[Dict[str, Any]],
    output_dir: Path,
    hash_bom: str,
) -> Dict[str, Path]:
    output_dir.mkdir(parents=True, exist_ok=True)
    resumen = resumen_totales(comparativas)

    data = {
        "generado_en": datetime.now().isoformat(timespec="seconds"),
        "bom_hash": hash_bom,
        "proyecto": raw_bom.get("proyecto"),
        "propietario": raw_bom.get("propietario"),
        "resumen": resumen,
        "comparativa": [c.to_dict() for c in comparativas],
    }

    paths = {
        "json": output_dir / "comparativa_precios.json",
        "csv": output_dir / "comparativa_precios.csv",
        "xlsx": output_dir / "comparativa_precios.xlsx",
        "html": output_dir / "cotizacion_recomendada.html",
        "tex": output_dir / "cotizacion_recomendada.tex",
        "md": output_dir / "resumen_cotizacion.md",
        "manifest": output_dir / "evidencias" / "manifest.json",
    }
    
    guardar_json(paths["json"], data)
    escribir_csv(paths["csv"], comparativas)
    escribir_xlsx(paths["xlsx"], raw_bom, materiales, comparativas, resumen, manifest)
    escribir_html(paths["html"], raw_bom, comparativas, resumen)
    escribir_latex(paths["tex"], raw_bom, comparativas, resumen)
    escribir_resumen_md(paths["md"], raw_bom, comparativas, resumen)
    guardar_json(paths["manifest"], manifest)
    return paths


def escribir_csv(path: Path, comparativas: List[ComparativaItem]) -> None:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        for r in comp.resultados_por_proveedor:
            rows.append({
                "material_id": m.id,
                "material": m.nombre_original,
                "normalizado": m.nombre_normalizado,
                "categoria": m.categoria,
                "unidad_bom": m.unidad,
                "cantidad_bom": m.cantidad,
                "proveedor": r.proveedor,
                "producto": r.producto,
                "marca": r.marca,
                "unidad_comercial": r.unidad_comercial,
                "contenido_comercial": r.contenido_comercial,
                "cantidad_comercial_a_comprar": r.cantidad_comercial_a_comprar,
                "precio_comercial": r.precio,
                "precio_equiv": r.precio_equivalente_unitario,
                "subtotal_real": r.subtotal_compra_real,
                "sobrante": r.sobrante,
                "tipo_coincidencia": r.tipo_coincidencia,
                "url": r.url,
                "disponibilidad": r.disponibilidad,
                "fecha_consulta": r.fecha_consulta,
                "metodo_extraccion": r.metodo_extraccion,
                "confianza": r.confianza,
                "score_total": r.score_total,
                "recomendado": bool(comp.resultado_recomendado and r is comp.resultado_recomendado),
                "observaciones": r.observaciones,
            })
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()) if rows else [])
        if rows:
            writer.writeheader()
            writer.writerows(rows)


def escribir_xlsx(
    path: Path,
    raw_bom: Dict[str, Any],
    materiales: List[MaterialBOM],
    comparativas: List[ComparativaItem],
    resumen: Dict[str, Any],
    manifest: List[Dict[str, Any]],
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except Exception as exc:
        (path.with_suffix(".xlsx.ERROR.txt")).write_text(str(exc), encoding="utf-8")
        return

    wb = openpyxl.Workbook()
    header_fill = PatternFill("solid", fgColor="0F766E")
    header_font = Font(color="FFFFFF", bold=True)

    def add_sheet(title: str, rows: List[Dict[str, Any]]):
        ws = wb.create_sheet(title)
        if not rows:
            ws.append(["sin datos"])
            return ws
        headers = list(rows[0].keys())
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in rows:
            ws.append([excel_scalar(row.get(h)) for h in headers])
        for col in ws.columns:
            letter = col[0].column_letter
            width = min(max(len(str(c.value or "")) for c in col) + 2, 60)
            ws.column_dimensions[letter].width = max(width, 10)
        return ws

    wb.remove(wb.active)
    bom_rows = []
    for idx, item in enumerate(raw_bom.get("materiales", raw_bom.get("items", [])), 1):
        if isinstance(item, dict):
            bom_rows.append({"n": idx, **item})
            
    add_sheet("BOM original", bom_rows)
    add_sheet("Materiales normalizados", [m.to_dict() for m in materiales])
    
    # Hojas por tienda
    add_sheet("Promart", filas_tienda_excel(comparativas, lambda p: p == "promart"))
    add_sheet("Sodimac_Maestro", filas_tienda_excel(comparativas, lambda p: p in {"sodimac", "maestro"}))
    add_sheet("MercadoLibre", filas_tienda_excel(comparativas, lambda p: p == "mercadolibre"))
    
    # Hojas generales
    add_sheet("Comparativa cruzada", filas_comparativa_cruzada(comparativas))
    add_sheet("Cotizacion_recomendada", filas_recomendadas_excel(comparativas))
    add_sheet("Plan_compra_consolidado", resumen.get("plan_compra_consolidado", []))
    add_sheet("Materiales_pendientes", filas_pendientes_excel(comparativas))
    
    add_sheet("Resumen por proveedor", [
        {
            "proveedor": k,
            "total_encontrado": v["total"],
            "items_con_precio": v["items_con_precio"],
            "items_sin_precio": v["items_sin_precio"],
            "cobertura_pct": v["cobertura_pct"],
            "estado": v["estado"],
            "advertencia": v["advertencia"],
        }
        for k, v in resumen.get("totales_por_proveedor", {}).items()
    ] + [{"proveedor": "RECOMENDADO MIXTO", "cobertura": resumen.get("cobertura_bom"), "compra_real_soles": resumen.get("total_compra_real_mixta")}])
    add_sheet("Diagnostico_proveedores", filas_diagnostico_proveedores(comparativas, manifest))
    add_sheet("Evidencias", manifest)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def excel_scalar(value: Any) -> Any:
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


def slug_resultado(resultado: ResultadoProveedor) -> str:
    if resultado.proveedor_slug:
        return resultado.proveedor_slug
    return resultado.proveedor.lower().replace(" ", "")


def filas_tienda_excel(comparativas: List[ComparativaItem], filter_func: Any) -> List[Dict[str, Any]]:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        candidatos = [
            r for r in comp.resultados_por_proveedor
            if filter_func(slug_resultado(r)) and resultado_aceptable_tienda(r)
        ]
        relevantes = [r for r in comp.resultados_por_proveedor if filter_func(slug_resultado(r))]
        if candidatos:
            mejor = max(candidatos, key=lambda r: r.score_total)
        else:
            mejor = max(
                relevantes,
                key=lambda r: (r.precio is not None, r.producto is not None, r.confianza),
                default=None,
            )
        rows.append({
            "material_id": m.id,
            "material_bom": m.nombre_original,
            "categoria": m.categoria,
            "cantidad_requerida": m.cantidad,
            "unidad_diseno": m.unidad,
            "query_usada": mejor.consulta if mejor else None,
            "producto_encontrado": mejor.producto if mejor else None,
            "precio": mejor.precio if mejor and resultado_aceptable_tienda(mejor) else None,
            "moneda": mejor.moneda if mejor else "PEN",
            "unidad_comercial": mejor.unidad_comercial if mejor else None,
            "contenido_comercial": mejor.contenido_comercial if mejor else None,
            "cantidad_comercial": mejor.cantidad_comercial_a_comprar if mejor else None,
            "precio_equivalente": mejor.precio_equivalente_unitario if mejor and resultado_aceptable_tienda(mejor) else None,
            "subtotal_real": mejor.subtotal_compra_real if mejor and resultado_aceptable_tienda(mejor) else None,
            "sobrante": mejor.sobrante if mejor else None,
            "url": mejor.url if mejor else None,
            "tipo_enlace": mejor.tipo_enlace if mejor else None,
            "metodo_extraccion": mejor.metodo_extraccion if mejor else None,
            "estado_proveedor": mejor.estado_proveedor if mejor else "pendiente_manual",
            "tipo_coincidencia": mejor.tipo_coincidencia if mejor else None,
            "confianza": mejor.score_total if mejor else 0.0,
            "observacion": (
                "; ".join(x for x in [mejor.observaciones, mejor.observacion_conversion, mejor.motivo_fallo] if x)
                if mejor else "Sin resultado registrado"
            ),
        })
    return rows


def filas_pendientes_excel(comparativas: List[ComparativaItem]) -> List[Dict[str, Any]]:
    return [
        {
            "material_id": comp.material_bom.id,
            "material": comp.material_bom.nombre_original,
            "cantidad": comp.material_bom.cantidad,
            "unidad": comp.material_bom.unidad,
            "motivo": comp.motivo_recomendacion,
        }
        for comp in comparativas
        if comp.resultado_recomendado is None
    ]


def filas_comparativa_cruzada(comparativas: List[ComparativaItem]) -> List[Dict[str, Any]]:
    def celda(comp: ComparativaItem, slugs: set[str]) -> Any:
        candidatos = [
            r for r in comp.resultados_por_proveedor
            if slug_resultado(r) in slugs and resultado_aceptable_tienda(r)
        ]
        if candidatos:
            return min(candidatos, key=lambda r: float(r.subtotal_compra_real or float("inf"))).subtotal_compra_real
        estados = [r.estado_proveedor for r in comp.resultados_por_proveedor if slug_resultado(r) in slugs]
        if "bloqueado" in estados:
            return "bloqueado"
        if "unidad_no_convertible" in estados:
            return "unidad no convertible"
        if "baja_confianza" in estados:
            return "baja confianza"
        return "sin precio / link busqueda"

    rows = []
    for comp in comparativas:
        recomendado = comp.resultado_recomendado
        rows.append({
            "material": comp.material_bom.nombre_original,
            "cantidad": comp.material_bom.cantidad,
            "unidad": comp.material_bom.unidad,
            "promart": celda(comp, {"promart"}),
            "sodimac_maestro": celda(comp, {"sodimac", "maestro"}),
            "mercado_libre": celda(comp, {"mercadolibre"}),
            "mejor_opcion": (
                f"{recomendado.proveedor}: S/ {recomendado.subtotal_compra_real:.2f}"
                if recomendado and recomendado.subtotal_compra_real is not None else "pendiente"
            ),
            "estado": recomendado.tipo_coincidencia if recomendado else "pendiente_revision",
        })
    return rows


def filas_diagnostico_proveedores(
    comparativas: List[ComparativaItem],
    manifest: List[Dict[str, Any]],
) -> List[Dict[str, Any]]:
    filas = [dict(x) for x in manifest if x.get("tipo") == "diagnostico_http"]
    if filas:
        return filas
    for comp in comparativas:
        por_proveedor: Dict[str, List[ResultadoProveedor]] = defaultdict(list)
        for resultado in comp.resultados_por_proveedor:
            por_proveedor[slug_resultado(resultado)].append(resultado)
        for proveedor, resultados in por_proveedor.items():
            representativo = max(
                resultados,
                key=lambda r: (resultado_aceptable_tienda(r), r.producto is not None, r.confianza),
            )
            filas.append({
                "material_id": comp.material_bom.id,
                "material": comp.material_bom.nombre_original,
                "proveedor": proveedor,
                "consulta": representativo.consulta,
                "url_solicitada": representativo.url_solicitada or representativo.url,
                "url_final": representativo.url_final or representativo.url,
                "codigo_http": representativo.codigo_http,
                "hubo_redirect": representativo.hubo_redirect,
                "metodo": representativo.metodo_extraccion,
                "resultados_crudos": max((r.resultados_crudos for r in resultados), default=0),
                "resultados_aceptados": sum(1 for r in resultados if resultado_aceptable_tienda(r)),
                "estado": representativo.estado_proveedor,
                "motivo": representativo.motivo_fallo or representativo.observaciones,
            })
    return filas


def filas_recomendadas_excel(comparativas: List[ComparativaItem]) -> List[Dict[str, Any]]:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        r = comp.resultado_recomendado
        if r:
            rows.append({
                "material_id": m.id,
                "material": m.nombre_original,
                "cantidad_req": m.cantidad,
                "unidad_bom": m.unidad,
                "proveedor": r.proveedor,
                "producto": r.producto,
                "unidad_comercial": r.unidad_comercial,
                "contenido_comercial": r.contenido_comercial,
                "cantidad_a_comprar": r.cantidad_comercial_a_comprar,
                "precio_comercial": r.precio,
                "precio_equiv": r.precio_equivalente_unitario,
                "subtotal_real": r.subtotal_compra_real,
                "sobrante": r.sobrante,
                "url": r.url,
                "confianza": r.score_total,
                "justificacion": comp.motivo_recomendacion,
            })
        else:
            rows.append({
                "material_id": m.id,
                "material": m.nombre_original,
                "cantidad_req": m.cantidad,
                "unidad_bom": m.unidad,
                "producto": "Sin cotizacion",
                "subtotal_real": None,
            })
    return rows


def filas_comparativa(comparativas: List[ComparativaItem]) -> List[Dict[str, Any]]:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        recomendado = comp.resultado_recomendado
        for r in comp.resultados_por_proveedor:
            rows.append({
                "material_id": m.id,
                "material": m.nombre_original,
                "cantidad_bom": m.cantidad,
                "unidad_bom": m.unidad,
                "proveedor": r.proveedor,
                "producto": r.producto,
                "marca": r.marca,
                "unidad_comercial": r.unidad_comercial,
                "contenido_comercial": r.contenido_comercial,
                "cantidad_comercial_a_comprar": r.cantidad_comercial_a_comprar,
                "precio_comercial": r.precio,
                "precio_equivalente": r.precio_equivalente_unitario,
                "subtotal_real": r.subtotal_compra_real,
                "sobrante": r.sobrante,
                "tipo_coincidencia": r.tipo_coincidencia,
                "url": r.url,
                "disponibilidad": r.disponibilidad,
                "metodo": r.metodo_extraccion,
                "confianza": r.confianza,
                "score_total": r.score_total,
                "recomendado": "si" if recomendado and r is recomendado else "no",
                "observaciones": r.observaciones,
            })
    return rows


def escribir_html(path: Path, raw_bom: Dict[str, Any], comparativas: List[ComparativaItem], resumen: Dict[str, Any]) -> None:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        rec = comp.resultado_recomendado
        proveedor = html.escape(rec.proveedor if rec else "Pendiente")
        producto = html.escape(rec.producto if rec and rec.producto else "Sin precio verificable")
        precio_equiv = f"S/ {rec.precio_equivalente_unitario:.2f}" if rec and rec.precio_equivalente_unitario is not None else "-"
        cant_com = f"{rec.cantidad_comercial_a_comprar:g} {rec.unidad_comercial}" if rec and rec.cantidad_comercial_a_comprar is not None else "-"
        subtotal = f"S/ {rec.subtotal_compra_real:.2f}" if rec and rec.subtotal_compra_real is not None else "-"
        url = f'<a href="{html.escape(rec.url or "")}">fuente</a>' if rec and rec.url else "-"
        coinc_lbl = f'<span class="badge">{rec.tipo_coincidencia}</span>' if rec else ""
        rows.append(f"""
        <tr>
          <td><strong>{html.escape(m.nombre_original)}</strong><br><small>{html.escape(m.nombre_normalizado)}</small></td>
          <td>{m.cantidad:g} {html.escape(m.unidad)}</td>
          <td>{proveedor}</td>
          <td>{producto} {coinc_lbl}</td>
          <td>{cant_com}</td>
          <td class="num">{precio_equiv}</td>
          <td class="num"><strong>{subtotal}</strong></td>
          <td>{url}</td>
          <td>{html.escape(comp.motivo_recomendacion)}</td>
        </tr>""")

    comp_blocks = []
    for comp in comparativas:
        m = comp.material_bom
        trs = []
        for r in comp.resultados_por_proveedor:
            p_eq = 'S/ %.2f' % r.precio_equivalente_unitario if r.precio_equivalente_unitario is not None else '-'
            subt = 'S/ %.2f' % r.subtotal_compra_real if r.subtotal_compra_real is not None else '-'
            precio_tienda_str = 'S/ %.2f' % r.precio if r.precio is not None else '-'
            trs.append(f"""
            <tr>
              <td>{html.escape(r.proveedor)}</td>
              <td>{html.escape(r.producto or 'No extraido')} <span class="badge">{r.tipo_coincidencia}</span></td>
              <td class="num">{precio_tienda_str} ({r.unidad_comercial or 'und'})</td>
              <td class="num">{p_eq}</td>
              <td class="num">{subt}</td>
              <td>{html.escape(r.metodo_extraccion)}</td>
              <td>{r.score_total:.2f}</td>
              <td>{'<a href=\"%s\">url</a>' % html.escape(r.url) if r.url else '-'}</td>
            </tr>""")
        comp_blocks.append(f"""
        <h3>{html.escape(m.nombre_original)}</h3>
        <table class="mini"><thead><tr><th>Proveedor</th><th>Producto</th><th>Precio Tienda</th><th>Precio Equiv</th><th>Subtotal Real</th><th>Metodo</th><th>Score</th><th>Fuente</th></tr></thead>
        <tbody>{''.join(trs)}</tbody></table>""")

    html_doc = f"""<!doctype html>
<html lang="es">
<head>
<meta charset="utf-8">
<title>Cotizacion recomendada multi-proveedor</title>
<style>
@page {{ size: A4; margin: 1.5cm; }}
body {{ font-family: Arial, sans-serif; color:#17212b; margin: 24px; }}
h1 {{ color:#0f766e; font-size:20px; margin-bottom:4px; }}
h2 {{ color:#0f766e; font-size:15px; margin-top:24px; }}
h3 {{ font-size:12px; margin:18px 0 6px; }}
.meta {{ color:#5c6670; font-size:12px; margin-bottom:14px; }}
table {{ width:100%; border-collapse: collapse; font-size:11px; margin:10px 0 18px; }}
th {{ background:#0f766e; color:white; text-align:left; padding:7px; }}
td {{ border-bottom:1px solid #e5e7eb; padding:6px; vertical-align:top; }}
.num {{ text-align:right; white-space:nowrap; }}
.totales {{ width:420px; margin-left:auto; }}
.mini {{ font-size:10px; }}
a {{ color:#1d4ed8; text-decoration:none; }}
.badge {{ background:#f3f4f6; color:#374151; padding:2px 4px; border-radius:3px; font-size:9px; }}
@media print {{ body {{ margin:0; }} }}
</style>
</head>
<body>
<h1>Cotizacion recomendada comercial mixta</h1>
<p class="meta">Proyecto: {html.escape(str(raw_bom.get('proyecto', '')))}<br>
Propietario: {html.escape(str(raw_bom.get('propietario', '')))}<br>
Generado: {datetime.now().isoformat(timespec='seconds')}</p>

<h2>Resumen</h2>
<table class="totales">
<tr><th>Concepto</th><th>Total</th></tr>
<tr><td>Total de compra real comercial mixta</td><td class="num"><strong>S/ {resumen['total_compra_real_mixta']:.2f}</strong></td></tr>
<tr><td>Total valor neto neto (BOM)</td><td class="num">S/ {resumen['total_recomendado_materiales']:.2f}</td></tr>
<tr><td>Cobertura del BOM</td><td class="num">{resumen['cobertura_bom']}</td></tr>
<tr><td>Materiales sin precio</td><td class="num">{len(resumen['materiales_sin_precio'])}</td></tr>
<tr><td>Materiales baja confianza</td><td class="num">{len(resumen['materiales_baja_confianza'])}</td></tr>
</table>

<h2>Mejor opcion por material (Compra Comercial Real)</h2>
<table>
<thead><tr><th>Material</th><th>Cant. BOM</th><th>Proveedor</th><th>Producto</th><th>Cant. Com.</th><th>P.Equiv.</th><th>Subtotal Real</th><th>URL</th><th>Justificacion</th></tr></thead>
<tbody>{''.join(rows)}</tbody>
</table>

<h2>Comparativa por proveedor</h2>
{''.join(comp_blocks)}
</body>
</html>"""
    path.write_text(html_doc, encoding="utf-8")


def escribir_latex(path: Path, raw_bom: Dict[str, Any], comparativas: List[ComparativaItem], resumen: Dict[str, Any]) -> None:
    rows = []
    for comp in comparativas:
        m = comp.material_bom
        r = comp.resultado_recomendado
        if r:
            cant_com_str = f"{r.cantidad_comercial_a_comprar:g} {r.unidad_comercial}"
            subt_real = r.subtotal_compra_real
            rows.append(
                f"{tex(m.nombre_original)} & {m.cantidad:g} {tex(m.unidad)} & "
                f"{tex(r.proveedor)} & {cant_com_str} & "
                f"{fmt_tex_price(r.precio_equivalente_unitario)} & "
                f"\\textbf{{{fmt_tex_price(subt_real)}}} \\\\"
            )
        else:
            rows.append(
                f"{tex(m.nombre_original)} & {m.cantidad:g} {tex(m.unidad)} & "
                f"\\textit{{Pendiente}} & -- & -- & -- \\\\"
            )
            
    content = f"""% Archivo generado por cotizador_multi_proveedor.py
% Requiere: \\usepackage{{booktabs,longtable,array}}
\\section{{Cotizacion multi-proveedor de materiales - Compra Comercial Real}}

La cotizacion fue generada con busqueda automatica multi-fase (exacta, comercial y equivalente superior) e integra conversiones a empaques reales de proveedores (rollos, tubos y cajas completas).

\\begin{{longtable}}{{p{{5.0cm}} p{{1.5cm}} p{{2.2cm}} p{{2.2cm}} r r}}
\\toprule
\\textbf{{Material}} & \\textbf{{Cant. BOM}} & \\textbf{{Proveedor}} & \\textbf{{Cant. Com.}} & \\textbf{{P.Equiv (S/)}} & \\textbf{{Subtotal Real (S/)}} \\\\
\\midrule
{chr(10).join(rows)}
\\midrule
\\multicolumn{{5}}{{r}}{{\\textbf{{Total de compra real comercial mixta (S/.)}}}} & \\textbf{{{resumen['total_compra_real_mixta']:.2f}}} \\\\
\\multicolumn{{5}}{{r}}{{Total valor neto equivalente (S/.)}} & {resumen['total_recomendado_materiales']:.2f} \\\\
\\bottomrule
\\end{{longtable}}

\\noindent\\textbf{{Cobertura del BOM:}} {resumen['cobertura_bom']}.\\\\
\\textbf{{Materiales sin precio verificable:}} {len(resumen['materiales_sin_precio'])}.\\\\
\\textbf{{Materiales con baja confianza:}} {len(resumen['materiales_baja_confianza'])}.
"""
    path.write_text(content, encoding="utf-8")


def tex(value: Any) -> str:
    s = str(value if value is not None else "")
    replacements = {
        "\\": r"\textbackslash{}",
        "&": r"\&",
        "%": r"\%",
        "$": r"\$",
        "#": r"\#",
        "_": r"\_",
        "{": r"\{",
        "}": r"\}",
    }
    for old, new in replacements.items():
        s = s.replace(old, new)
    return s


def fmt_tex_price(value: Optional[float]) -> str:
    return "--" if value is None else f"{value:.2f}"


def escribir_resumen_md(path: Path, raw_bom: Dict[str, Any], comparativas: List[ComparativaItem], resumen: Dict[str, Any]) -> None:
    lines = [
        "# Resumen de Cotización Multi-Proveedor",
        "",
        f"**Proyecto:** {raw_bom.get('proyecto', 'N/E')}",
        f"**Propietario:** {raw_bom.get('propietario', 'N/E')}",
        f"**Fecha de generación:** {datetime.now().isoformat(timespec='seconds')}",
        "",
        "## Resumen de Totales",
        "",
        f"- **Total de Compra Real Mixta (unidades comerciales):** S/ {resumen.get('total_compra_real_mixta', 0.0):.2f}",
        f"- **Total Recomendado Neto (valores de diseño):** S/ {resumen.get('total_recomendado_materiales', 0.0):.2f}",
        f"- **Cobertura total del BOM:** {resumen.get('cobertura_bom', '0/0')}",
        f"- **Materiales sin precio:** {len(resumen['materiales_sin_precio'])}",
        f"- **Materiales con baja confianza:** {len(resumen['materiales_baja_confianza'])}",
        f"- **Calidad de la comparativa:** {resumen.get('mensaje_calidad', '')}",
        "",
        "## Totales por Tienda (Compra Comercial Real)",
        "",
    ]
    
    for prov, data in resumen.get("totales_por_proveedor", {}).items():
        total_texto = f"S/ {data['total']:.2f}" if data["total"] is not None else "sin precio total"
        lines.append(
            f"- **{prov.upper()}**: {total_texto} "
            f"({data['items_con_precio']} encontrados, {data['items_sin_precio']} pendientes, "
            f"cobertura {data['cobertura_pct']:.1f} %, estado {data['estado']}). "
            f"{data['advertencia']}"
        )
        
    lines.extend([
        "",
        "## Tabla Cruzada Comparativa de Precios",
        "",
        "Esta tabla presenta el costo total de compra real en cada tienda (considerando empaques de rollos, tubos y cantidades comerciales) versus la mejor opción mixta recomendada.",
        "",
        "| Material | Cant. | Unidad | Promart | Sodimac/Maestro | MercadoLibre | Mejor Opción | Observación |",
        "|---|---|---|---|---|---|---|---|",
    ])
    
    for comp in comparativas:
        mat = comp.material_bom
        
        # Buscar precios para cada tienda
        def get_subtotal_str(filter_func):
            cands = [
                x for x in comp.resultados_por_proveedor
                if filter_func(slug_resultado(x)) and resultado_aceptable_tienda(x)
            ]
            if cands:
                mejor = max(cands, key=lambda x: x.score_total)
                if mejor.score_tecnico >= 0.45:
                    return f"S/ {mejor.subtotal_compra_real:.2f}"
            return "-"
            
        p_promart = get_subtotal_str(lambda p: p == "promart")
        p_sodimac = get_subtotal_str(lambda p: p in {"sodimac", "maestro"})
        p_ml = get_subtotal_str(lambda p: p == "mercadolibre")
        
        rec = comp.resultado_recomendado
        if rec:
            mejor_opcion = f"{rec.proveedor.upper()}: S/ {rec.subtotal_compra_real:.2f}"
            obs = f"[{rec.tipo_coincidencia.upper()}] {rec.observaciones or ''}"
            if rec.observacion_conversion:
                obs += f"; {rec.observacion_conversion}"
        else:
            mejor_opcion = "*Pendiente*"
            obs = comp.motivo_recomendacion
            
        lines.append(
            f"| {mat.nombre_original} | {mat.cantidad:g} | {mat.unidad} | {p_promart} | {p_sodimac} | {p_ml} | {mejor_opcion} | {obs} |"
        )
        
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def escribir_reportes_adicionales(
    output_dir: Path,
    comparativas: List[ComparativaItem],
    resumen: Dict[str, Any],
    manifest: Optional[List[Dict[str, Any]]] = None,
    raw_bom: Optional[Dict[str, Any]] = None,
) -> None:
    # 1. Crear directorios
    output_dir.mkdir(parents=True, exist_ok=True)
    (output_dir / "evidencias").mkdir(parents=True, exist_ok=True)
    (output_dir / "por_tienda").mkdir(parents=True, exist_ok=True)
    
    # 2. Generar por_tienda/promart.md, sodimac_maestro.md, mercado_libre.md
    grupos_tienda = {
        "promart": {
            "file": "promart.md",
            "title": "Promart",
            "slugs": {"promart"},
            "filter": lambda p: p == "promart"
        },
        "sodimac_maestro": {
            "file": "sodimac_maestro.md",
            "title": "Sodimac / Maestro",
            "slugs": {"sodimac", "maestro"},
            "filter": lambda p: p in {"sodimac", "maestro"}
        },
        "mercado_libre": {
            "file": "mercado_libre.md",
            "title": "Mercado Libre Perú",
            "slugs": {"mercadolibre"},
            "filter": lambda p: p == "mercadolibre"
        }
    }
    
    for key, info in grupos_tienda.items():
        generar_md_tienda(
            output_dir / "por_tienda" / info["file"],
            info["title"],
            info["filter"],
            comparativas,
            info["slugs"],
        )
        
    # 4. Generar por_tienda/recomendada_mixta.md
    generar_md_recomendada_mixta(output_dir / "por_tienda" / "recomendada_mixta.md", comparativas, resumen)
    
    # 5. Generar evidencias/enlaces_productos.md
    generar_md_enlaces_evidencias(output_dir / "evidencias" / "enlaces_productos.md", comparativas)
    
    # 6. Generar pendientes_revision.md
    generar_md_pendientes(output_dir / "pendientes_revision.md", comparativas)
    generar_diagnostico_proveedores(
        output_dir / "diagnostico_proveedores.md",
        comparativas,
        manifest or [],
        resumen,
    )
    
    # 7. Generar README.md
    generar_md_readme(output_dir / "README.md", raw_bom, resumen)
    if raw_bom is not None:
        generar_informe_latex_imprimible(output_dir, raw_bom, comparativas, resumen)


def generar_diagnostico_proveedores(
    path: Path,
    comparativas: List[ComparativaItem],
    manifest: List[Dict[str, Any]],
    resumen: Dict[str, Any],
) -> None:
    aceptados: Dict[Tuple[str, str], int] = defaultdict(int)
    for comp in comparativas:
        for resultado in comp.resultados_por_proveedor:
            if resultado_aceptable_tienda(resultado):
                aceptados[(slug_resultado(resultado), resultado.consulta)] += 1

    lines = [
        "# Diagnostico de proveedores",
        "",
        f"**Estado global:** {resumen.get('mensaje_calidad', '')}",
        "",
        "| Proveedor | Material probado | URL | Estado HTTP | Redirect | Metodo | Resultados crudos | Resultados aceptados | Estado | Motivo de fallo |",
        "|---|---|---|---:|---|---|---:|---:|---|---|",
    ]
    intentos = filas_diagnostico_proveedores(comparativas, manifest)
    for intento in intentos:
        proveedor = intento.get("proveedor", "")
        consulta = intento.get("consulta", "")
        url = intento.get("url_solicitada") or intento.get("url_final") or ""
        url_md = f"[abrir]({url})" if url else "-"
        lines.append(
            "| {proveedor} | {material} | {url} | {http} | {redirect} | {metodo} | {crudos} | {aceptados} | {estado} | {motivo} |".format(
                proveedor=proveedor,
                material=intento.get("material") or consulta,
                url=url_md,
                http=intento.get("codigo_http") if intento.get("codigo_http") is not None else "-",
                redirect="si" if intento.get("hubo_redirect") else "no",
                metodo=intento.get("metodo", ""),
                crudos=intento.get("resultados_crudos", 0),
                aceptados=intento.get("resultados_aceptados", aceptados.get((proveedor, consulta), 0)),
                estado=intento.get("estado", ""),
                motivo=(intento.get("motivo") or "-").replace("|", "/"),
            )
        )

    lines.extend([
        "",
        "## Hallazgos tecnicos",
        "",
        "- **Promart:** expone un endpoint publico VTEX con productos, precios, stock y enlaces directos.",
        "- **Sodimac:** la URL historica redirige a la portada; el fallback usa el catalogo unificado Falabella y conserva solo ofertas cuyo vendedor es SODIMAC.",
        "- **Maestro:** el dominio historico puede agotar conexion; el fallback consulta el catalogo unificado y conserva solo ofertas identificadas con vendedor MAESTRO.",
        "- **Mercado Libre:** la API publica puede responder 403 y el HTML puede mostrar verificacion de trafico. No se intentan saltar controles; se conserva el enlace de busqueda verificable.",
        "- Los resultados con precio pero categoria, especificacion o unidad no confiable se mantienen como evidencia y no compiten por la recomendacion.",
    ])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _descripcion_compra(resultado: ResultadoProveedor) -> str:
    cantidad = resultado.cantidad_comercial_a_comprar
    unidad = resultado.unidad_comercial or "unidad"
    contenido = resultado.contenido_comercial
    unidad_contenido = resultado.unidad_contenido or ""
    if cantidad is None:
        return "--"
    detalle = f"{cantidad:g} {unidad}"
    if contenido and contenido != 1:
        detalle += f" de {contenido:g} {unidad_contenido}"
    return detalle


def generar_informe_latex_imprimible(
    output_dir: Path,
    raw_bom: Dict[str, Any],
    comparativas: List[ComparativaItem],
    resumen: Dict[str, Any],
) -> None:
    informe_dir = output_dir / "informe-final"
    informe_dir.mkdir(parents=True, exist_ok=True)
    tex_path = informe_dir / "informe_comparativo_proveedores.tex"

    plan_consolidado = resumen.get("plan_compra_consolidado", [])
    distribucion: Dict[str, Dict[str, float]] = defaultdict(lambda: {"items": 0.0, "total": 0.0})
    for fila in plan_consolidado:
        if fila.get("proveedor"):
            prov = fila["proveedor"]
            distribucion[prov]["items"] += 1
            distribucion[prov]["total"] += float(fila.get("subtotal_compra_real") or 0.0)
    proveedor_principal = max(distribucion, key=lambda p: distribucion[p]["items"], default="ninguno")
    mejor_cobertura = max(
        resumen.get("totales_por_proveedor", {}),
        key=lambda p: resumen["totales_por_proveedor"][p]["cobertura_pct"],
        default="ninguno",
    )
    if proveedor_principal == mejor_cobertura:
        conclusion = (
            f"Se recomienda realizar un pedido mixto. {proveedor_principal} concentra la mayor "
            "cantidad de partidas seleccionadas y ofrece la mayor cobertura individual; Promart "
            "debe utilizarse para las partidas donde reduce el costo real consolidado. Los materiales "
            "pendientes deben confirmarse antes de emitir la orden final."
        )
    else:
        conclusion = (
            f"Se recomienda realizar un pedido mixto siguiendo la tabla de compra recomendada. "
            f"{proveedor_principal} concentra la mayor cantidad de partidas seleccionadas, mientras que "
            f"{mejor_cobertura} ofrece la mayor cobertura individual. Los materiales pendientes deben "
            "confirmarse antes de emitir la orden final."
        )

    proveedor_rows = []
    for proveedor, datos in resumen.get("totales_por_proveedor", {}).items():
        total = "--" if datos["total"] is None else f"S/ {datos['total']:.2f}"
        proveedor_rows.append(
            f"{tex(proveedor)} & {datos['items_con_precio']}/{len(comparativas)} & "
            f"{datos['cobertura_pct']:.1f}\\% & {tex(total)} & {tex(datos['estado'])} \\\\"
        )

    recomendadas_rows = []
    pendientes_rows = []
    for fila in plan_consolidado:
        if fila.get("proveedor"):
            enlace = f"\\href{{\\detokenize{{{fila['url']}}}}}{{verificar}}" if fila.get("url") else "--"
            compra = f"{fila['cantidad_comercial_a_comprar']:g} {fila['unidad_comercial'] or 'unidad'}"
            if fila.get("contenido_comercial") and fila["contenido_comercial"] != 1:
                compra += f" de {fila['contenido_comercial']:g} {fila.get('unidad_contenido') or ''}"
            recomendadas_rows.append(
                f"{tex(fila['descripcion_representativa'])} ({fila['partidas_bom']} partidas) & "
                f"{fila['cantidad_total_bom']:g} {tex(fila['unidad_bom'])} & "
                f"{tex(fila['proveedor'])} & {tex(fila['producto'] or '')} & "
                f"{tex(compra)} & S/ {fila['precio_comercial']:.2f} & "
                f"S/ {fila['subtotal_compra_real']:.2f} & {tex(fila['tipo_coincidencia'])} & {enlace} \\\\"
            )
        else:
            pendientes_rows.append(
                f"{tex(fila['descripcion_representativa'])} & {fila['cantidad_total_bom']:g} {tex(fila['unidad_bom'])} & "
                f"{tex(fila['observacion'])} \\\\"
            )

    proveedor_detalle = []
    grupos = [
        ("Promart", {"promart"}),
        ("Sodimac / Maestro", {"sodimac", "maestro"}),
        ("Mercado Libre Peru", {"mercadolibre"}),
    ]
    for titulo, slugs in grupos:
        filas = filas_tienda_excel(comparativas, lambda slug, allowed=slugs: slug in allowed)
        detalle_rows = []
        for fila in filas:
            precio = "--" if fila["precio"] is None else f"S/ {fila['precio']:.2f}"
            subtotal = "--" if fila["subtotal_real"] is None else f"S/ {fila['subtotal_real']:.2f}"
            detalle_rows.append(
                f"{tex(fila['material_bom'])} & {tex(fila['producto_encontrado'] or '--')} & "
                f"{tex(precio)} & {tex(subtotal)} & {tex(fila['estado_proveedor'])} \\\\"
            )
        proveedor_detalle.append(
            f"\\section{{{tex(titulo)}}}\n"
            "\\begin{longtable}{p{0.23\\textwidth} p{0.35\\textwidth} r r p{0.13\\textwidth}}\n"
            "\\toprule Material & Producto & Precio & Subtotal real & Estado \\\\ \\midrule\\endhead\n"
            + "\n".join(detalle_rows)
            + "\n\\bottomrule\\end{longtable}\n"
        )

    contenido = rf"""\documentclass[11pt,a4paper]{{article}}
\usepackage[utf8]{{inputenc}}
\usepackage[T1]{{fontenc}}
\usepackage[spanish]{{babel}}
\usepackage[a4paper,margin=1.6cm]{{geometry}}
\usepackage{{longtable,booktabs,array,xcolor,hyperref,xurl,pdflscape}}
\hypersetup{{colorlinks=true,urlcolor=blue,linkcolor=black}}
\setlength{{\parindent}}{{0pt}}
\setlength{{\parskip}}{{5pt}}
\renewcommand{{\arraystretch}}{{1.16}}
\title{{Informe comparativo de cotizacion de materiales electricos}}
\author{{{tex(raw_bom.get('propietario', 'Por completar'))}}}
\date{{{tex(datetime.now().strftime('%d/%m/%Y'))}}}
\begin{{document}}
\maketitle
\thispagestyle{{empty}}
\vfill
\begin{{center}}
\large Informe listo para revision, impresion y gestion de compra\\[8pt]
\normalsize {tex(raw_bom.get('proyecto', 'Proyecto de instalaciones electricas'))}
\end{{center}}
\vfill
\newpage
\tableofcontents
\newpage

\section{{Resumen ejecutivo}}
Se evaluaron {len(comparativas)} materiales del BOM sin alterar sus unidades de diseno. La cobertura de la recomendacion mixta es {tex(resumen.get('cobertura_bom', '0/0'))} y el costo de compra real, considerando rollos, tubos, paquetes y unidades completas, es \textbf{{S/ {resumen.get('total_compra_real_mixta', 0.0):.2f}}}.

El informe separa precio faltante de precio cero, documenta bloqueos y exige URL y confianza tecnica para recomendar.

\textbf{{Estado de calidad:}} {tex(resumen.get('mensaje_calidad', ''))}

\section{{Metodologia}}
\begin{{itemize}}
\item Se preserva la cantidad y unidad original del BOM.
\item Para rollos, tubos y paquetes se calcula la cantidad entera a comprar con redondeo hacia arriba.
\item El subtotal recomendado es el costo comercial real, no una multiplicacion fraccionaria imposible.
\item Solo compiten ofertas con URL, precio, categoria compatible y conversion confiable.
\item Los superiores inmediatos se identifican y se marcan para revision; no se acepta una capacidad inferior.
\item No se evaden captchas ni verificaciones de trafico.
\end{{itemize}}

\section{{Cobertura y totales por proveedor}}
\begin{{tabular}}{{lccrl}}
\toprule Proveedor & Partidas & Cobertura & Total parcial & Estado \\ \midrule
{chr(10).join(proveedor_rows)}
\bottomrule
\end{{tabular}}

\textit{{Los totales parciales de proveedores con faltantes no son presupuestos completos y no deben compararse como si cubrieran el mismo alcance.}}

\section{{Cotizacion recomendada mixta}}
\begin{{landscape}}
\small
\begin{{longtable}}{{p{{0.19\linewidth}} p{{0.07\linewidth}} p{{0.09\linewidth}} p{{0.24\linewidth}} p{{0.12\linewidth}} r r p{{0.09\linewidth}} c}}
\toprule Material & Requerido & Proveedor & Producto & Compra comercial & Precio & Subtotal & Coincidencia & Fuente \\ \midrule\endhead
{chr(10).join(recomendadas_rows)}
\bottomrule
\end{{longtable}}
\normalsize
\end{{landscape}}

\section{{Analisis y decision de compra}}
{tex(conclusion)}

\begin{{itemize}}
{chr(10).join(f"\\item {tex(prov)}: {int(datos['items'])} partidas recomendadas, S/ {datos['total']:.2f}." for prov, datos in sorted(distribucion.items()))}
\end{{itemize}}

\textbf{{Decision propuesta:}} emitir ordenes separadas segun la cotizacion mixta, priorizando el proveedor indicado para cada material. No consolidar todo en una sola tienda mientras su cobertura sea parcial o existan pendientes tecnicos.

\section{{Materiales pendientes de confirmacion}}
\begin{{longtable}}{{p{{0.35\textwidth}} p{{0.12\textwidth}} p{{0.48\textwidth}}}}
\toprule Material & Cantidad & Motivo / accion \\ \midrule\endhead
{chr(10).join(pendientes_rows) if pendientes_rows else r'\multicolumn{3}{c}{No quedaron materiales pendientes.} \\'}
\bottomrule
\end{{longtable}}

\section{{Detalle por proveedor}}
{chr(10).join(proveedor_detalle)}

\section{{Conclusiones}}
\begin{{enumerate}}
\item La comparacion debe basarse en cobertura y subtotal real de compra, no en ceros para precios ausentes.
\item La conversion comercial evita rechazar cables por rollo y tuberias por tramo cuando el contenido es verificable.
\item El pedido recomendado es mixto y debe seguir las partidas de la tabla principal.
\item Los enlaces deben revisarse el dia de la compra porque precio y stock pueden cambiar.
\item Las partidas pendientes requieren confirmacion tecnica o precio manual con URL antes de cerrar el presupuesto.
\end{{enumerate}}

\end{{document}}
"""
    tex_path.write_text(contenido, encoding="utf-8")

    readme = informe_dir / "README.md"
    readme.write_text(
        "# Informe final imprimible\n\n"
        "- `informe_comparativo_proveedores.tex`: fuente LaTeX.\n"
        "- `informe_comparativo_proveedores.pdf`: informe compilado listo para imprimir.\n",
        encoding="utf-8",
    )
    compilar_informe_latex(tex_path)


def compilar_informe_latex(tex_path: Path) -> None:
    import subprocess

    if shutil.which("pdflatex") is None:
        (tex_path.parent / "COMPILACION_PENDIENTE.txt").write_text(
            "pdflatex no esta instalado. Compile dos veces el archivo .tex.",
            encoding="utf-8",
        )
        return
    for _ in range(2):
        proceso = subprocess.run(
            ["pdflatex", "-interaction=nonstopmode", "-halt-on-error", tex_path.name],
            cwd=tex_path.parent,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            errors="replace",
            check=False,
        )
        if proceso.returncode != 0:
            (tex_path.parent / "COMPILACION_ERROR.txt").write_text(proceso.stdout, encoding="utf-8")
            return
    for extension in ("aux", "log", "out", "toc"):
        auxiliar = tex_path.with_suffix(f".{extension}")
        if auxiliar.exists():
            auxiliar.unlink()


def generar_md_tienda(
    path: Path,
    title: str,
    filter_func: Any,
    comparativas: List[ComparativaItem],
    slugs: set[str],
) -> None:
    lines = [
        f"# Reporte de Cotización - {title}",
        "",
        "| Material BOM | Producto Encontrado | Cant. Req. | Und. Dis. | Und. Com. | Cont. Com. | Cant. Com. a Comprar | Precio Com. | Precio Equiv. | Subtotal Real | Sobrante | URL | Conf. | Observación |",
        "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
    ]
    
    encontrados = 0
    pendientes = 0
    total_compra_real = sum(
        float(fila["subtotal_compra_real"] or 0.0)
        for fila in construir_plan_compra_consolidado(comparativas, slugs)
        if fila["subtotal_compra_real"] is not None
    )
    materiales_pendientes = []
    
    for comp in comparativas:
        mat = comp.material_bom
        candidatos = [
            r for r in comp.resultados_por_proveedor
            if filter_func(slug_resultado(r)) and resultado_aceptable_tienda(r)
        ]
        if candidatos:
            mejor = max(candidatos, key=lambda r: r.score_total)
            if mejor.score_tecnico >= 0.45 and mejor.metodo_extraccion != "fallback":
                encontrados += 1
                prod_name = mejor.producto or ""
                cant_req = f"{mat.cantidad:g}"
                und_dis = mat.unidad
                und_com = mejor.unidad_comercial or ""
                cont_com = f"{mejor.contenido_comercial:g}"
                cant_com = f"{mejor.cantidad_comercial_a_comprar:g}" if mejor.cantidad_comercial_a_comprar else ""
                precio_com = f"S/ {mejor.precio:.2f}" if mejor.precio is not None else ""
                precio_eq = f"S/ {mejor.precio_equivalente_unitario:.2f}" if mejor.precio_equivalente_unitario is not None else ""
                subtotal = f"S/ {mejor.subtotal_compra_real:.2f}" if mejor.subtotal_compra_real is not None else ""
                sobrante = f"{mejor.sobrante:g} {mejor.unidad_contenido or ''}" if mejor.sobrante else "0"
                url_link = f"[Enlace]({mejor.url})" if mejor.url else "-"
                conf = f"{mejor.score_total:.2f}"
                obs = mejor.observaciones or ""
                if mejor.observacion_conversion:
                    obs = f"{obs}; {mejor.observacion_conversion}"
                
                lines.append(
                    f"| {mat.nombre_original} | {prod_name} | {cant_req} | {und_dis} | {und_com} | {cont_com} | {cant_com} | {precio_com} | {precio_eq} | {subtotal} | {sobrante} | {url_link} | {conf} | {obs} |"
                )
                continue
                
        pendientes += 1
        materiales_pendientes.append(mat.nombre_original)
        lines.append(
            f"| {mat.nombre_original} | *No encontrado / confianza baja* | {mat.cantidad:g} | {mat.unidad} | - | - | - | - | - | - | - | - | - | - |"
        )
        
    cobertura = f"{encontrados}/{len(comparativas)}"
    
    summary_lines = [
        "",
        "## Resumen por Tienda",
        "",
        f"- **Cobertura del BOM:** {cobertura}",
        f"- **Materiales encontrados:** {encontrados}",
        f"- **Materiales pendientes:** {pendientes}",
        f"- **Total de compra real consolidado con unidades comerciales:** S/ {total_compra_real:.2f}",
        "",
    ]
    if materiales_pendientes:
        summary_lines.append("### Materiales Pendientes")
        for m in materiales_pendientes:
            summary_lines.append(f"- {m}")
        summary_lines.append("")
        
    lines[2:2] = summary_lines
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generar_md_recomendada_mixta(path: Path, comparativas: List[ComparativaItem], resumen: Dict[str, Any]) -> None:
    lines = [
        "# Reporte de Cotización Recomendada Mixta",
        "",
        "## Resumen General",
        "",
        f"- **Total Recomendado Mixto (Compra Real):** S/ {resumen.get('total_compra_real_mixta', 0.0):.2f}",
        f"- **Total Recomendado Materiales (Valor Neto Equivalente):** S/ {resumen.get('total_recomendado_materiales', 0.0):.2f}",
        f"- **Cobertura del BOM:** {resumen['cobertura_bom']}",
        f"- **Materiales sin precio:** {len(resumen['materiales_sin_precio'])}",
        f"- **Materiales con baja confianza:** {len(resumen['materiales_baja_confianza'])}",
        "",
        "| Material consolidado | Partidas BOM | Producto Recomendado | Proveedor | Cant. Req. | Unidad BOM | Cant. Com. a Comprar | Unidad Com. | Precio Com. | Precio Equiv. | Subtotal Real | Sobrante | Tipo Coincidencia | URL |",
        "|---|---|---|---|---|---|---|---|---|---|---|---|---|---|",
    ]
    for fila in resumen.get("plan_compra_consolidado", []):
        if fila.get("proveedor"):
            lines.append(
                f"| {fila['descripcion_representativa']} | {fila['partidas_bom']} | {fila['producto']} | {fila['proveedor']} | "
                f"{fila['cantidad_total_bom']:g} | {fila['unidad_bom']} | {fila['cantidad_comercial_a_comprar']:g} | "
                f"{fila['unidad_comercial']} | S/ {fila['precio_comercial']:.2f} | S/ {fila['precio_equivalente_unitario']:.2f} | "
                f"S/ {fila['subtotal_compra_real']:.2f} | {fila['sobrante']:g} {fila.get('unidad_contenido') or ''} | "
                f"{fila['tipo_coincidencia']} | [Enlace]({fila['url']}) |"
            )
        else:
            lines.append(
                f"| {fila['descripcion_representativa']} | {fila['partidas_bom']} | *Pendiente de revisión* | - | "
                f"{fila['cantidad_total_bom']:g} | {fila['unidad_bom']} | - | - | - | - | - | - | pendiente | - |"
            )
            
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generar_md_enlaces_evidencias(path: Path, comparativas: List[ComparativaItem]) -> None:
    lines = [
        "# Enlaces de Verificación de Productos",
        "",
        "| Material BOM | Proveedor | Producto | Precio | Unidad Comercial | Fecha consulta | Tipo Coincidencia | Estado | Confianza | Observación | URL |",
        "|---|---|---|---|---|---|---|---|---|---|---|",
    ]
    
    for comp in comparativas:
        mat = comp.material_bom
        for r in comp.resultados_por_proveedor:
            precio_str = f"S/ {r.precio:.2f}" if r.precio is not None else "sin precio"
            etiqueta = "Ver producto" if r.tipo_enlace == "producto" else "Link busqueda"
            url_link = f"[{etiqueta}]({r.url})" if r.url else "Manual/Pendiente"
            obs = r.observaciones or ""
            if r.observacion_conversion:
                obs = f"{obs}; {r.observacion_conversion}"
            
            lines.append(
                f"| {mat.nombre_original} | {r.proveedor} | {r.producto or 'sin producto'} | {precio_str} | {r.unidad_comercial or ''} | {r.fecha_consulta or ''} | {r.tipo_coincidencia} | {r.estado_proveedor} | {r.score_total:.2f} | {obs} | {url_link} |"
            )
            
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generar_md_pendientes(path: Path, comparativas: List[ComparativaItem]) -> None:
    categorizado = {
        "no_encontrado": [],
        "no_convertible": [],
        "baja_confianza": [],
        "similar_revision": [],
        "bloqueo_datos": [],
        "falta_regla": []
    }
    
    for comp in comparativas:
        mat = comp.material_bom
        r = comp.resultado_recomendado
        
        if not r:
            has_prices = any(x.precio is not None for x in comp.resultados_por_proveedor)
            if not has_prices:
                estados = {x.estado_proveedor for x in comp.resultados_por_proveedor}
                if estados & {"bloqueado", "timeout", "redirect_no_util", "parser_fallo"}:
                    detalle = ", ".join(sorted(estados & {"bloqueado", "timeout", "redirect_no_util", "parser_fallo"}))
                    categorizado["bloqueo_datos"].append((mat, f"Proveedores sin datos utiles: {detalle}.", "revisar link / completar precio manual verificable"))
                else:
                    categorizado["no_encontrado"].append((mat, "No se encontraron productos en ningun proveedor.", "confirmar especificacion tecnica / revisar link"))
            else:
                non_convertible = any(x.precio is not None and getattr(x, "confianza_conversion", 1.0) == 0.0 for x in comp.resultados_por_proveedor)
                if non_convertible:
                    categorizado["no_convertible"].append((mat, "Unidades comerciales encontradas no son convertibles de forma segura.", "agregar regla de conversion"))
                else:
                    categorizado["baja_confianza"].append((mat, "Coincidencia tecnica o total por debajo del umbral minimo (0.45).", "completar precio manual / buscar equivalente superior"))
        else:
            if r.tipo_coincidencia == "similar_pendiente_revision":
                categorizado["similar_revision"].append((mat, f"Producto similar encontrado ({r.producto}), pero requiere revision manual.", "confirmar especificacion tecnica"))
            elif r.score_tecnico < 0.5 and r.tipo_coincidencia == "equivalente_superior":
                categorizado["similar_revision"].append((mat, f"Equivalente superior encontrado ({r.producto}), requiere aprobacion de seccion/diametro.", "confirmar especificacion tecnica"))
                
    lines = [
        "# Reporte de Materiales Pendientes de Revisión",
        "",
        "Este reporte lista los materiales del BOM que no pudieron cotizarse automáticamente de forma confiable, agrupados por su causa, sugiriendo una acción correctiva.",
        "",
    ]
    
    causas = [
        ("no_encontrado", "1. No encontrado en proveedores", "No se encontraron precios en la web para el término buscado."),
        ("no_convertible", "2. Encontrado pero unidad comercial no convertible", "Los productos encontrados usan unidades comerciales que no tienen regla de conversión confiable en `unidades_comerciales.json`."),
        ("baja_confianza", "3. Encontrado pero confianza baja", "Los productos se encontraron pero su score técnico o total es inferior a 0.45."),
        ("similar_revision", "4. Producto similar, requiere revisión", "Se encontró un producto similar o equivalente superior, pero por seguridad requiere aprobación manual."),
        ("bloqueo_datos", "5. Proveedor bloqueó o no expuso datos", "El scraping falló o devolvió errores de red/bloqueo."),
        ("falta_regla", "6. Falta regla de equivalencia", "No existe regla para buscar equivalentes de esta categoría.")
    ]
    
    for key, title, desc in causas:
        items = categorizado[key]
        lines.append(f"## {title}")
        lines.append(desc)
        lines.append("")
        if not items:
            lines.append("*Ningún material en esta sección.*")
        else:
            lines.append("| Material | Cantidad Requerida | Detalle | Acción Propuesta |")
            lines.append("|---|---|---|---|")
            for mat, detalle, accion in items:
                lines.append(f"| {mat.nombre_original} | {mat.cantidad:g} {mat.unidad} | {detalle} | **{accion}** |")
        lines.append("")
        
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def generar_md_readme(path: Path, raw_bom: Dict[str, Any], resumen: Dict[str, Any]) -> None:
    lines = [
        f"# Cotización de Materiales - {raw_bom.get('proyecto', 'Proyecto')}",
        "",
        "Este directorio contiene los entregables de la cotización multi-proveedor del proyecto.",
        "",
        "## Estructura de Salidas",
        "",
        "- `comparativa_precios.json`: Base de datos completa en JSON.",
        "- `comparativa_precios.csv`: Tabla de comparativas en formato CSV.",
        "- `comparativa_precios.xlsx`: Hoja de cálculo Excel organizada con pestañas por tienda.",
        "- `cotizacion_recomendada.html`: Cotización recomendada en HTML interactivo.",
        "- `cotizacion_recomendada.tex`: Tabla LaTeX lista para insertar en el capítulo II del expediente.",
        "- `resumen_cotizacion.md`: Resumen general de la cotización y totales.",
        "- `pendientes_revision.md`: Reporte detallado de materiales que requieren validación manual.",
        "- `evidencias/`: Carpeta con evidencias de scraping y enlaces de verificación.",
        "  - `enlaces_productos.md`: Listado ordenado con los enlaces directos a las tiendas.",
        "- `por_tienda/`: Carpeta con los presupuestos individuales de cada tienda.",
        "  - `promart.md`: Presupuesto y cobertura en Promart.",
        "  - `sodimac_maestro.md`: Presupuesto y cobertura en Sodimac/Maestro.",
        "  - `mercado_libre.md`: Presupuesto y cobertura en Mercado Libre.",
        "  - `recomendada_mixta.md`: Presupuesto optimizado mixto.",
        "",
        "## Resumen de la Cotización Mixta Recomendada",
        "",
        f"- **Total de compra real mixta (unidades comerciales):** S/ {resumen.get('total_compra_real_mixta', 0.0):.2f}",
        f"- **Total recomendado neto de materiales (BOM):** S/ {resumen.get('total_recomendado_materiales', 0.0):.2f}",
        f"- **Cobertura del BOM:** {resumen.get('cobertura_bom')}",
        f"- **Materiales sin precio:** {len(resumen['materiales_sin_precio'])}",
        f"- **Materiales baja confianza:** {len(resumen['materiales_baja_confianza'])}",
    ]
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def validar_proveedores(slugs: Iterable[str]) -> List[str]:
    out = []
    for slug in slugs:
        slug = slug.strip()
        if not slug:
            continue
        if slug not in PROVEEDORES_DISPONIBLES:
            raise ValueError(f"Proveedor no soportado: {slug}")
        out.append(slug)
    return out


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Cotizador automatico multi-proveedor")
    p.add_argument("--bom", required=True, help="Ruta al BOM JSON")
    p.add_argument("--proveedores", default=",".join(DEFAULT_PROVEEDORES), help="Lista separada por comas")
    p.add_argument("--output", default="build/cotizacion", help="Directorio de salida")
    p.add_argument("--output-proyecto", help="Copia adicional organizada dentro de un proyecto")
    p.add_argument("--max-resultados", type=int, default=5, help="Maximo de resultados por proveedor")
    p.add_argument("--max-materiales", type=int, help="Limitar materiales procesados")
    p.add_argument("--usar-cache", action="store_true", default=True, help="Usar cache local")
    p.add_argument("--no-cache", action="store_false", dest="usar_cache", help="No usar cache local")
    p.add_argument("--refrescar-cache", action="store_true", help="Ignorar cache de lectura y actualizarla")
    p.add_argument("--offline", action="store_true", help="No consultar internet")
    p.add_argument("--usar-fixtures", action="store_true", help="Usar HTML de pruebas en modo offline")
    p.add_argument("--timeout", type=float, default=10.0, help="Timeout HTTP por proveedor")
    p.add_argument("--delay", type=float, default=1.0, help="Rate limit entre consultas por proveedor")
    p.add_argument("--debug-proveedor", choices=sorted(PROVEEDORES_DISPONIBLES), help="Ejecutar solo un proveedor")
    p.add_argument("--debug-material", help="Ejecutar solo el material que coincida con este texto")
    p.add_argument(
        "--permitir-estimados",
        action="store_true",
        help="Generar presupuesto academico completo, marcando de forma explicita los precios estimados",
    )
    return p.parse_args()


def main() -> int:
    args = parse_args()
    bom_path = Path(args.bom)
    output_dir = Path(args.output)
    if not bom_path.exists():
        print(f"Error: no existe BOM {bom_path}", file=sys.stderr)
        return 2
        
    proveedores_slugs = validar_proveedores(args.proveedores.split(","))
    if args.debug_proveedor:
        proveedores_slugs = [args.debug_proveedor]
    raw_bom, materiales = cargar_bom(bom_path, args.max_materiales)
    if args.debug_material:
        objetivo = normalizar_nombre_material(args.debug_material)
        candidatos = [
            material for material in materiales
            if objetivo in material.nombre_normalizado
            or material.nombre_normalizado in objetivo
            or args.debug_material.lower() in material.nombre_original.lower()
        ]
        if not candidatos:
            print(f"Error: no se encontro material para debug: {args.debug_material}", file=sys.stderr)
            return 2
        materiales = candidatos[:1]
    hash_b = bom_hash(bom_path)
    cache = cargar_cache(CACHE_PATH)

    print(f"BOM: {bom_path} ({len(materiales)} materiales, hash {hash_b})")
    print(f"Proveedores: {', '.join(proveedores_slugs)}")
    print(f"Salida primaria: {output_dir}")
    if args.output_proyecto:
        print(f"Salida adicional del proyecto: {args.output_proyecto}")
    if args.offline:
        print("Modo offline activo")

    # Limpiar evidencias previas en directorio de salida
    evidencias_previas = output_dir / "evidencias"
    if evidencias_previas.exists():
        shutil.rmtree(evidencias_previas)

    comparativas, manifest = construir_comparativa(
        materiales, proveedores_slugs, args, cache, hash_b, output_dir
    )
    if args.usar_cache and not args.offline:
        guardar_cache(cache, CACHE_PATH)

    paths = generar_salidas(raw_bom, materiales, comparativas, manifest, output_dir, hash_b)
    resumen = resumen_totales(comparativas)
    escribir_reportes_adicionales(output_dir, comparativas, resumen, manifest, raw_bom)
    entrega_final_primaria = None
    if args.permitir_estimados:
        from entrega_academica import generar_entrega_final

        entrega_final_primaria = generar_entrega_final(
            raw_bom, comparativas, output_dir, resultado_aceptable_tienda
        )
    
    # Generar una copia adicional si se solicita.
    if args.output_proyecto:
        proj_dir = Path(args.output_proyecto)
        if proj_dir.exists():
            # Limpiar evidencias previas
            shutil.rmtree(proj_dir, ignore_errors=True)
        proj_dir.mkdir(parents=True, exist_ok=True)
        generar_salidas(raw_bom, materiales, comparativas, manifest, proj_dir, hash_b)
        escribir_reportes_adicionales(proj_dir, comparativas, resumen, manifest, raw_bom)
        if args.permitir_estimados:
            from entrega_academica import generar_entrega_final

            generar_entrega_final(raw_bom, comparativas, proj_dir, resultado_aceptable_tienda)

    print("\nArchivos generados en salida primaria:")
    for key, path in paths.items():
        print(f"  {key}: {path}")
        
    print(f"\nTotal recomendado comercial mixta (Compra Real): S/ {resumen['total_compra_real_mixta']:.2f}")
    print(f"Total recomendado materials (Valor Neto): S/ {resumen['total_recomendado_materiales']:.2f}")
    print(f"Cobertura BOM: {resumen['cobertura_bom']}")
    print(f"Materiales sin precio: {len(resumen['materiales_sin_precio'])}")
    print(f"Materiales baja confianza: {len(resumen['materiales_baja_confianza'])}")
    if entrega_final_primaria:
        data_final = entrega_final_primaria["data"]
        print(f"Presupuesto academico completo: S/ {data_final['total_recomendado_mixto']:.2f}")
        print(f"Grupos de compra consolidados: {data_final['grupos_compra']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
